# -*- coding: utf-8 -*-
"""ExcelWriterHandler class, used to safely open and close files and apply styles.

@author  : Donald Erb
Created on Dec 9, 2020

Notes
-----
openpyxl is imported within methods of ExcelWriterHandler in order
to reduce the import time of mcetl since the writer is only needed
if writing to Excel.

"""


from collections import defaultdict
import copy
from pathlib import Path
import traceback

import pandas as pd
import PySimpleGUI as sg

from . import utils
#openpyxl is imported within methods of ExcelWriterHandler


class ExcelWriterHandler:
    """
    A helper for pandas's ExcelWriter for opening/saving files and applying styles.

    This class is used for ensuring that an existing file is closed before
    saving and/or writing, if appending, and that all desired styles are
    ready for usage before writing to Excel. Styles can either be openpyxl
    NamedStyle objects in order to make the style available in the Excel
    workbook, or dictionaries detailing the style, such as
    {'font': Font(...), 'border': Border(...), ...}.

    Parameters
    ----------
    file_name : str or Path
        The file name or path for the Excel file to be created.
    new_file : bool, optional
        If False (default), will append to an existing file. If True, or if
        no file currently exists with the given file_name, a new file will
        be created, even if a file with the same name currently exists.
    styles : dict(str, None or dict or str or openpyxl.styles.named_styles.NamedStyle)
        A dictionary of either nested dictionaries used to create openpyxl
        style objects (Alignment, Font, etc.), a string indicating the name
        of an openpyxl NamedStyle to use, a NamedStyle object or None (will use
        the default style if None). All styles in the dictionary will be added
        to the Excel workbook if they do not currently exist in the workbook.
        See Examples below to see various valid inputs for styles.
    writer : pd.ExcelWriter or None
        The ExcelWriter (_OpenpyxlWriter from pandas) used for writing
        to Excel. If it is a pd.ExcelWriter, its engine must be "openpyxl".
    **kwargs
        Any additional keyword arguments to pass to pd.ExcelWriter.

    Attributes
    ----------
    styles : dict(str, dict)
        A nested dictionary of dictionaries, used to create openpyxl
        NamedStyle objects to include in self.writer.book. The styles
        are used as a class attribute to ensure that the necessary
        styles are always included in the Excel book.
    style_cache : dict(str, tuple(str, str or openpyxl.styles.cell_style.StyleArray))
        The currently implemented styles within the Excel workbook. Used
        to quickly apply styles to cells without having to constanly set
        all of the cell attributes (cell.font, cell.fill, etc.). The dictionary
        value is a tuple of the cell attribute name and the value to set.
        Will either be ('style', string indicating NamedStyle.name)
        if using NamedStyles or ('_style', openpyxl StyleArray) to indicate
        an anonomous style. The call to set the cell attribute for a
        desired key would be setattr(cell, *style_cache[key]).
    writer : pd.ExcelWriter
        The ExcelWriter (_OpenpyxlWriter from pandas) used for writing
        to Excel.

    Notes
    -----
    Either file_name or writer must be specified at initialization.

    Examples
    --------
    Below is a partial example of various allowable input styles. Can be openpyxl
    NamedStyle, str, None, or dictionary. (Note that NamedStyle, Font, Border, and Side
    are gotten by importing from openpyxl.styles)

    >>> styles = {
            # Would make the style 'Even Header' available in the output Excel file
            'fitting_header_even': NamedStyle(
                name='Even Header',
                font=Font(size=12, bold=True),
                border=Border(bottom=Side(style='thin')),
                number_format='0.0'
            ),
            # Would use same format as 'fitting_header_even'
            'fitting_header_odd': 'Even Header',
            # Basically just replaces NamedStyle from 'fitting_header_even' with
            # dict and removes the 'name' key. A new style would not be created
            # in the output Excel file.
            'fitting_subheader_even': dict(
                font=Font(size=12, bold=True),
                aligment=Aligment(bottom=Side(style='thin')),
                number_format='0.0'
            ),
            # Same as 'fitting_subheader_even', but doesn't require importing
            # from openpyxl. Basically just replaces all openpyxl objects with dict.
            'fitting_subheader_odd': dict(
                font=dict(size=12, bold=True),
                aligment=dict(bottom=dict(style='thin')),
                number_format='0.0'
            ),
            # Same as 'fitting_subheader_odd', but will create a NamedStyle (and
            # add the style to the Excel file) since 'name' is within the dictionary.
            'fitting_columns_even': dict(
                name='New Style',
                font=dict(size=12, bold=True),
                aligment=dict(bottom=dict(style='thin')),
                number_format='0.0'
            ),
            # Would use default style set by openpyxl
            'fitting_columns_odd': {},
            # Would also use default style
            'fitting_descriptors_even': None
        }

    """

    styles = {
        'fitting_header_even': {
            'font': dict(size=12, bold=True),
            'fill': dict(fill_type='solid', start_color='F9B381', end_color='F9B381'),
            'border': dict(bottom=dict(style='thin')),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True)
        },
        'fitting_header_odd': {
            'font': dict(size=12, bold=True),
            'fill': dict(fill_type='solid', start_color='73A2DB', end_color='73A2DB'),
            'border': dict(bottom=dict(style='thin')),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True)
        },
        'fitting_subheader_even': {
            'font': dict(bold=True),
            'fill': dict(fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'),
            'border': dict(bottom=dict(style='thin')),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True)
        },
        'fitting_subheader_odd': {
            'font': dict(bold=True),
            'fill': dict(fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'),
            'border': dict(bottom=dict(style='thin')),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True)
        },
        'fitting_columns_even': {
            'fill': dict(fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'),
            'alignment': dict(horizontal='center', vertical='center'),
            'number_format': '0.00'
        },
        'fitting_columns_odd': {
            'fill': dict(fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'),
            'alignment': dict(horizontal='center', vertical='center'),
            'number_format': '0.00'
        },
        'fitting_descriptors_even': {
            'font': dict(bold=True),
            'fill': dict(fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True),
            'number_format': '0.000'
        },
        'fitting_descriptors_odd': {
            'font': dict(bold=True),
            'fill': dict(fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True),
            'number_format': '0.000'
        }
    }

    def __init__(self, file_name=None, new_file=False, styles=None, writer=None, **kwargs):
        """
        Raises
        ------
        TypeError
            Raised if both file_name and writer are None.
        ValueError
            Raised if the engine of the input writer is not "openpyxl".

        """

        if file_name is None and writer is None:
            raise TypeError(
                'Both file_name and writer cannot be None when creating an ExcelWriterHandler.'
            )
        elif writer is None:
            self.writer = self._create_writer(file_name, new_file, **kwargs)
        elif writer.engine != 'openpyxl':
            raise ValueError(
                'ExcelWriterHandler can only use ExcelWriters with "openpyxl" engine.'
            )
        else:
            self.writer = writer

        self.style_cache = {}
        input_styles = styles if styles is not None else {}
        total_styles = {}
        # use dict.fromkeys to preserve ordering
        for key in dict.fromkeys((*input_styles.keys(), *self.styles.keys())).keys():
            if key in input_styles:
                value = input_styles[key]
            else:
                value = self.styles[key]
            total_styles[key] = value

        self.add_styles(total_styles)


    def __str__(self):
        return f'{self.__class__.__name__}(path={self.writer.path})'


    def _create_writer(self, file_name, new_file, **kwargs):
        """
        Creates the pandas ExcelWriter.

        Parameters
        ----------
        file_name : str or Path
            The file name or path for the Excel file to be created.
        new_file : bool
            If False, will append to an existing file. If True, or if
            no file currently exists with the given file_name, a new file will
            be created, even if a file with the same name currently exists.
        **kwargs
            Any additional keyword arguments to pass to pd.ExcelWriter.

        Returns
        -------
        writer : pd.ExcelWriter
            The ExcelWriter with the correct mode set.

        Notes
        -----
        If appending to a file and the file is open, warns the user that any
        unsaved or future changes to the file after creating the ExcelWriter
        will be lost.

        """

        path = Path(file_name)
        if new_file or not path.exists():
            mode = 'w'
        else:
            mode = 'a'
            try:
                #TODO: this only errors if file is currently open in Windows, need to find other solution
                # that is os-independent
                path.rename(path)
            except PermissionError:
                sg.popup_ok(
                    (f'{path.name} is about to be loaded in Python.\n\nTo keep '
                     'any current unsaved changes, save the file before closing '
                     'this window.\n\nAny changes to the file made within Excel '
                     'until the file is saved in Python will be lost.\n'),
                    title='Close File', icon=utils._LOGO
                )

        # TODO switch this to logging later, and make it log for either mode
        if mode == 'a':
            print(f'Loading {file_name}. May take a while...')
        writer = pd.ExcelWriter(file_name, engine='openpyxl', mode=mode, **kwargs)
        if mode == 'a':
            print(f'Done loading {file_name}.')

        return writer


    def save_excel_file(self):
        """
        Tries to save the Excel file, and handles any PermissionErrors.

        Saving can be cancelled if other changes to self.writer are desired
        before saving, or if saving is no longer desired (the file must be
        open while trying to save to allow cancelling the save).

        """

        path = Path(self.writer.path)
        # Ensures that the folder destination exists
        path.parent.mkdir(parents=True, exist_ok=True)
        print('Saving Excel file. May take a while...')  # TODO switch to logging later
        while True:
            try:
                self.writer.save()
                print('\nSaved Excel file.')  # TODO switch to logging later
                break

            except PermissionError:
                window = sg.Window(
                    'Save Error',
                    layout=[
                        [sg.Text((f'Trying to overwrite {path.name}.\n\n'
                                'Please close the file and press Proceed'
                                ' to save.\nPress Discard to not save.\n'))],
                        [sg.Button('Discard'),
                        sg.Button('Proceed', button_color=utils.PROCEED_COLOR)]
                    ],
                    icon=utils._LOGO
                )
                response = window.read()[0]
                window.close()
                window = None
                if response == 'Discard':
                    break


    def add_styles(self, styles):
        """
        Adds styles to the Excel workbook and update self.style_cache.

        Parameters
        ----------
        styles : dict(str, None or dict or str or openpyxl.styles.named_styles.NamedStyle)
            A dictionary of either nested dictionaries used to create openpyxl
            style objects (Alignment, Font, etc.), a string indicating the name
            of an openpyxl NamedStyle to use, a NamedStyle object or None (will use
            the default style if None). All styles in the dictionary will be added
            to the Excel workbook if they do not currently exist in the workbook.

        Notes
        -----
        The ordering of items within styles will be preserved, so that
        if two NamedStyles are input with the same name, the one appearing
        first in the dictionary will be created, and the second will be made
        to refer to the first.

        """

        new_styles = {}
        for key, value in styles.items():
            if isinstance(value, dict):
                style = self._create_openpyxl_objects(value)
            else:
                style = value
            new_styles[key] = style

        self._update_style_cache(new_styles)


    def _update_style_cache(self, styles):
        """
        Updates self.style_cache with new styles.

        Parameters
        ----------
        styles : dict(str, None or dict or str or openpyxl.styles.named_styles.NamedStyle)
            A dictionary of either None, nested dictionaries containing openpyxl
            style objects (eg. {'font': Font(...), 'border': Border(...)}),
            a string indicating the name of an openpyxl NamedStyle to use,
            or a NamedStyle object. All styles in the dictionary will be added
            to the self.writer.book if they do not currently exist in the workbook.

        Raises
        ------
        ValueError
            Raised if a string is given as a style but there is no NamedStyle
            within the workbook whose name matches the string.

        """

        from openpyxl.styles import NamedStyle

        temp_sheet = self.writer.book.create_sheet('temp_sheet')
        temp_cell = temp_sheet['A1']
        default_style = copy.copy(temp_cell._style) # should just be None

        # delayed_styles marks the NamedStyle names for any styles that are not
        # currently in the workbook
        delayed_styles = defaultdict(list)
        for key, style in styles.items():
            if isinstance(style, NamedStyle):
                if style.name not in self.writer.book.named_styles:
                    self.writer.book.add_named_style(style)
                self.style_cache[key] = ('style', style.name)
            elif isinstance(style, str):
                if style in self.writer.book.named_styles:
                    self.style_cache[key] = ('style', style)
                else:
                    delayed_styles[style].append(key)
            elif style is None:
                self.style_cache[key] = ('_style', default_style)
            else:
                for style_attribute, values in style.items():
                    setattr(temp_cell, style_attribute, values)
                self.style_cache[key] = ('_style', copy.copy(temp_cell._style))
                # reset back to default style each time to prevent styles
                # from overlapping if not all attributes are used
                temp_cell._style = default_style

        self.writer.book.remove(temp_sheet)

        if delayed_styles:
            for (cell_attribute, style_reference) in tuple(self.style_cache.values()):
                if style_reference in delayed_styles:
                    for cache_key in delayed_styles[style_reference]:
                        self.style_cache[cache_key] = (cell_attribute, style_reference)
                    delayed_styles.pop(style_reference)

            if delayed_styles:
                raise ValueError((
                    'The following NamedStyles need to be created '
                    f'before usage: {list(delayed_styles.keys())}.'
                ))

    @classmethod
    def _create_openpyxl_objects(cls, style):
        """
        Creates openpyxl objects from the input style.

        If 'name' is in the input style, will create a NamedStyle; otherwise,
        will create all of the necessary openpyxl style objects (Font, Alignment, etc.).

        Parameters
        ----------
        style : dict(str, str or dict or openpyxl style object (Font, Border, etc.))
            Will convert a dictionary of values into the necessary keyword
            arguments to create a openpyxl objects (Font, Alignment, etc.).

        Returns
        -------
        dict or openpyxl.styles.named_styles.NamedStyle
            If 'name' is within the input style dictionary, then returns a
            NamedStyle; otherwise, returns a dictionary containing openpyxl
            style objects (Font, Alignment, etc.).

        """
        from openpyxl.styles import NamedStyle

        kwargs = {}
        for key in ('alignment', 'border', 'fill', 'font', 'number_format', 'protection'):
            if key in style:
                kwargs[key] = getattr(cls, f'_openpyxl_{key}')(style[key])

        if 'name' in style:
            return NamedStyle(style['name'], **kwargs)
        else:
            return kwargs


    @classmethod
    def _openpyxl_number_format(cls, values):
        """
        Ensures that the input number_format is valid for openpyxl.

        Parameters
        ----------
        values : str or None
            The desired number_format.

        Returns
        -------
        str
            If the input values was None, returns '', otherwise,
            returns the input values.

        Raises
        ------
        TypeError
            Raised if the input is not None or a string.

        """
        if values is not None and not isinstance(values, str):
            raise TypeError('number_format must be a string or None.')

        return values if values is not None else ''


    @classmethod
    def _openpyxl_protection(cls, values):
        """
        Creates an openpyxl.styles.Protection object.

        Parameters
        ----------
        values : dict or openpyxl.styles.Protection
            Either a dictionary of keyword arguments to create an openpyxl
            Protection, or an openpyxl Protection.

        Returns
        -------
        openpyxl.styles.Protection
            The Protection object.

        """
        from openpyxl.styles import Protection

        if isinstance(values, Protection):
            return values
        else:
            return Protection(**values)


    @classmethod
    def _openpyxl_alignment(cls, values):
        """
        Creates an openpyxl.styles.Alignment object.

        Parameters
        ----------
        values : dict or openpyxl.styles.Alignment
            Either a dictionary of keyword arguments to create an openpyxl
            Alignment, or an openpyxl Alignment.

        Returns
        -------
        openpyxl.styles.Alignment
            The Alignment object.

        """
        from openpyxl.styles import Alignment

        if isinstance(values, Alignment):
            return values
        else:
            return Alignment(**values)


    @classmethod
    def _openpyxl_color(cls, values):
        """
        Creates an openpyxl.styles.Color object.

        Parameters
        ----------
        values : str or dict or openpyxl.styles.Color
            Either a strig or dictionary of keyword arguments to create
            an openpyxl Color, or an openpyxl Color.

        Returns
        -------
        openpyxl.styles.Color
            The Color object.

        """
        from openpyxl.styles import Color

        if isinstance(values, Color):
            return values
        elif isinstance(values, str):
            return Color(values)
        else:
            return Color(**values)


    @classmethod
    def _openpyxl_font(cls, values):
        """
        Creates an openpyxl.styles.Font object.

        Parameters
        ----------
        values : str or dict or openpyxl.styles.Font
            Either a strig or dictionary of keyword arguments to create
            an openpyxl Font, or an openpyxl Font.

        Returns
        -------
        openpyxl.styles.Font
            The Font object.

        """
        from openpyxl.styles import Font

        if isinstance(values, Font):
            return values
        elif isinstance(values, str):
            return Font(values)

        kwargs = {}
        for key, value in values.items():
            if key == 'color':
                kwargs[key] = cls._openpyxl_color(value)
            else:
                kwargs[key] = value

        return Font(**kwargs)


    @classmethod
    def _openpyxl_side(cls, values):
        """
        Creates an openpyxl.styles.Side object.

        Parameters
        ----------
        values : str or dict or openpyxl.styles.Side
            Either a strig or dictionary of keyword arguments to create
            an openpyxl Side, or an openpyxl Side.

        Returns
        -------
        openpyxl.styles.Side
            The Side object.

        """
        from openpyxl.styles import Side

        if isinstance(values, Side):
            return values
        elif isinstance(values, str):
            return Side(values)

        kwargs = {}
        for key, value in values.items():
            if key == 'color':
                kwargs[key] = cls._openpyxl_color(value)
            else:
                kwargs[key] = value

        return Side(**kwargs)


    @classmethod
    def _openpyxl_border(cls, values):
        """
        Creates an openpyxl.styles.Border object.

        Parameters
        ----------
        values : dict or openpyxl.styles.Border
            Either a dictionary of keyword arguments to create an openpyxl
            Border, or an openpyxl Border.

        Returns
        -------
        openpyxl.styles.Border
            The Border object.

        """
        from openpyxl.styles import Border

        if isinstance(values, Border):
            return values

        kwargs = {}
        for key, value in values.items():
            if key in ('left', 'right', 'top', 'bottom', 'diagonal'):
                kwargs[key] = cls._openpyxl_side(value)
            else:
                kwargs[key] = value

        return Border(**kwargs)


    @classmethod
    def _openpyxl_fill(cls, values):
        """
        Creates an openpyxl.styles.Fill object.

        Parameters
        ----------
        values : dict or openpyxl.styles.Fill
            Either a dictionary of keyword arguments to create an openpyxl
            PatternFill or GradientFill, or an openpyxl Fill.

        Returns
        -------
        openpyxl.styles.Fill
            The Fill object. Either a PatternFill or a GradientFill

        """
        from openpyxl.styles import Fill, GradientFill, PatternFill

        if isinstance(values, Fill):
            return values

        gradient_kwargs = {}
        pattern_kwargs = {}
        for key, value in values.items():
            gradient_key = pattern_key = val = None
            if key in ('fgColor', 'bgColor', 'start_color', 'end_color'):
                pattern_key = key
                val = cls._openpyxl_color(value)
            if key in ('patternType', 'fill_type'):
                pattern_key = key
            if key in ('type', 'fill_type'):
                gradient_key = 'type' # GradientFill does not take fill_type key
            if key == 'stop':
                gradient_key = key
                val = [cls._openpyxl_color(v) for v in value]
            if val is None:
                val = value

            if pattern_key:
                pattern_kwargs[pattern_key] = val
            if gradient_key:
                gradient_kwargs[gradient_key] = val
            if pattern_key is None and gradient_key is None:
                # all PatternFill keys should be covered above, but assign
                # anyway to cover unforseen differences in openpyxl versions
                pattern_kwargs[key] = val
                gradient_kwargs[key] = val

        try:
            output = PatternFill(**pattern_kwargs)
        except (TypeError, ValueError):
            output = GradientFill(**gradient_kwargs)

        return output


    @classmethod
    def test_excel_styles(cls, styles):
        """
        Tests whether the input styles create valid Excel styles using openpyxl.

        Parameters
        ----------
        styles : dict(str, None or dict or str or openpyxl.styles.named_styles.NamedStyle)
            The dictionary of styles to test. Values in the dictionary can
            either be None, a nested dictionary with the necessary keys and values
            to create an openpyxl NamedStyle, a string (which would refer to another
            NamedStyle.name), or openpyxl.styles.NamedStyle objects.

        Returns
        -------
        success : bool
            Returns True if all input styles successfully create openpyxl
            NamedStyle objects; otherwise, returns False.

        """

        from openpyxl.styles import NamedStyle

        failed_styles = []
        for key, style in styles.items():
            try:
                if isinstance(style, dict):
                    style_kwargs = cls._create_openpyxl_objects(style)
                    if isinstance(style_kwargs, dict):
                        # convert to NamedStyle to check the number_format
                        NamedStyle(**style_kwargs)
                elif style is not None and not isinstance(style, (str, NamedStyle)):
                    raise TypeError((
                        'The style must be None, a dictionary, a NamedStyle, '
                        'or a string that refers to a NamedStyle.name.'
                    ))
            except:
                failed_styles.append((key, traceback.format_exc()))

        if failed_styles:
            success = False
            print('The following input styles were incorrect:\n')
            for failure, traceback_message in failed_styles:
                print(f'\nKey: {failure}\n{traceback_message}')
        else:
            success = True
            print('All input styles were successful.\n')

        return success
