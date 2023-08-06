# -*- coding: utf-8 -*-
"""Provides GUIs to import data depending on the data source used, process and/or fit the data, and save everything to Excel.

@author: Donald Erb
Created on May 5, 2020

Notes
-----
The imports for the fitting and plotting guis are within their respective
functions to reduce the time it takes for this module to be imported. Likewise,
openpyxl is imported within _write_to_excel.

Attributes
----------
SAVE_FOLDER : pathlib.Path
    The file path to the folder in which all 'previous_files_{DataSource.name}.json'
    files are saved. Depends on operating system.

"""


import itertools
import json
import os
from pathlib import Path
import sys
import textwrap
import traceback
import warnings

import PySimpleGUI as sg

from . import utils
from .data_source import DataSource
from .excel_writer import ExcelWriterHandler
from .file_organizer import file_finder, file_mover, manual_file_finder
# openpyxl is imported within _write_to_excel


# the prefix of the filename used for saving previous files
_FILE_PREFIX = 'previous_files_'


def _get_save_location():
    """
    Gets the filepath for saving the previous files depending on the operating system.

    Returns
    -------
    pathlib.Path
        The absolute path to where the previous files json will be saved.

    Notes
    -----
    Tries to use environmental variables before using default locations, and
    tries to follow standard conventions. See the following links (and the
    additional links in the links) for more information:

    https://stackoverflow.com/questions/1024114/location-of-ini-config-files-in-linux-unix,
    https://specifications.freedesktop.org/basedir-spec/latest/

    """

    path = None
    if sys.platform.startswith('win'): # Windows
        path = Path(os.environ.get('LOCALAPPDATA') or '~/AppData/Local').joinpath('mcetl')
    elif sys.platform.startswith('darwin'): # Mac
        path = Path('~/Library/Application Support/mcetl')
    elif sys.platform.startswith(('linux', 'freebsd')): # Linux
        path = Path(os.environ.get('XDG_DATA_HOME') or '~/.local/share').joinpath('mcetl')

    if path is not None:
        try:
            if not path.expanduser().parent.is_dir():
                path = None
        except PermissionError:
            # permission is denied in the desired folder; will not really help
            # accessing, but allows this function to not fail so that user can
            # manually set SAVE_FOLDER before using launch_main_gui
            path = None

    if path is None:
        # unspecified os, the Windows/Mac/Linux places were wrong, or access denied
        path = Path('~/.mcetl')

    return path.expanduser()


SAVE_FOLDER = _get_save_location()


def _write_to_excel(dataframes, data_source, labels,
                    excel_writer_handler, plot_excel, plot_options):
    """
    Creates an Excel sheet from data within a list of dataframes.

    Parameters
    ----------
    dataframes : list(pd.DataFrame)
        A list of dataframes. Each dataframe contains all the raw data to
        put on one sheet in Excel.
    data_source : DataSource
        The selected DataSource.
    labels : list(dict)
        A list of dictionaries containing all of the sheet names, sample names,
        and subheader names. Each dictionary is for one dataset/Excel sheet.
        Relevant keys are 'sheet_name', 'sample_names', 'column_names'.
    excel_writer_handler : mcetl.excel_writer.ExcelWriterHandler
        The ExcelWriterHandler that contains the pandas ExcelWriter object for
        writing to Excel, and the styles for styling the cells in Excel.
    plot_excel : bool
        If True, will create a simple plot in Excel using the data_source's
        x_plot_index and y_plot_index.
    plot_options : list(dict)
        A list of dictionaries with values used to create the Excel plot
        if plot_excel is True.

    """

    from openpyxl.chart import Reference, Series, ScatterChart
    from openpyxl.chart.series import SeriesLabel, StrRef
    from openpyxl.utils.dataframe import dataframe_to_rows

    excel_writer = excel_writer_handler.writer
    style_cache = excel_writer_handler.style_cache

    # openpyxl uses 1-based indices
    first_row = data_source.excel_row_offset + 1
    first_column = data_source.excel_column_offset + 1

    for i, dataset in enumerate(dataframes):
        # Ensures that the sheet name is unique so it does not overwrite data;
        # not needed for openpyxl, but just a precaution
        current_sheets = [sheet.title.lower() for sheet in excel_writer.book.worksheets]
        sheet_name = labels[i]['sheet_name']
        sheet_base = sheet_name
        num = 1
        while sheet_name.lower() in current_sheets:
            sheet_name = f'{sheet_base}_{num}'
            num += 1

        worksheet = excel_writer.book.create_sheet(sheet_name)
        # Header values and formatting
        for j, header in enumerate(labels[i]['sample_names']):
            suffix = 'even' if j % 2 == 0 else 'odd'
            worksheet.merge_cells(
                start_row=first_row,
                start_column=first_column + sum(sum(entry) for entry in data_source.lengths[i][:j]),
                end_row=first_row,
                end_column=first_column + sum(sum(entry) for entry in data_source.lengths[i][:j + 1]) - 1
            )
            worksheet.cell(
                row=first_row,
                column=first_column + sum(sum(entry) for entry in data_source.lengths[i][:j]),
                value=header
            )
            for col in range(
                first_column + sum(sum(entry) for entry in data_source.lengths[i][:j]),
                first_column + sum(sum(entry) for entry in data_source.lengths[i][:j + 1])
            ):
                setattr(
                    worksheet.cell(row=first_row, column=col),
                    *style_cache['header_' + suffix]
                )
        # Subheader values and formatting
        flattened_lengths = list(itertools.chain.from_iterable(data_source.lengths[i]))
        subheaders = itertools.chain(labels[i]['column_names'], itertools.cycle(['']))
        for j, entry in enumerate(flattened_lengths):
            suffix = 'even' if j % 2 == 0 else 'odd'
            for col_index in range(entry):
                setattr(
                    worksheet.cell(
                        row=first_row + 1,
                        column=first_column + col_index + sum(flattened_lengths[:j]),
                        value=next(subheaders)
                    ),
                    *style_cache['subheader_' + suffix]
                )

        # Dataset values and formatting
        rows = dataframe_to_rows(dataset, index=False, header=False)
        for row_index, row in enumerate(rows, first_row + 2):
            entry = 1
            suffix = 'even'
            cycle = itertools.cycle(['odd', 'even'])
            for column_index, value in enumerate(row, first_column):
                if (column_index + 1 - first_column) > sum(flattened_lengths[:entry]):
                    suffix = next(cycle)
                    entry += 1
                setattr(
                    worksheet.cell(row=row_index, column=column_index, value=value),
                    *style_cache['columns_' + suffix]
                )

        worksheet.row_dimensions[first_row].height = 18
        worksheet.row_dimensions[first_row + 1].height = 30

        if plot_excel:
            x_min = plot_options[i]['x_min']
            x_max = plot_options[i]['x_max']
            y_min = plot_options[i]['y_min']
            y_max = plot_options[i]['y_max']
            last_row = len(dataset) + 1 + first_row

            # Prevents an error in Excel if using log scale and specified values are <= 0
            if plot_options[i]['x_log_scale']:
                if x_min is not None and x_min <= 0:
                    x_min = None
                if x_max is not None and x_max <= 0:
                    x_max = None
            if plot_options[i]['y_log_scale']:
                if y_min is not None and y_min <= 0:
                    y_min = None
                if y_max is not None and y_max <= 0:
                    y_max = None

            # Reverses x or y axes if min > max
            if None not in (x_min, x_max) and x_min > x_max:
                x_reverse = True
                x_min, x_max = x_max, x_min
            else:
                x_reverse = False

            if None not in (y_min, y_max) and y_min > y_max:
                y_reverse = True
                y_min, y_max = y_max, y_min
            else:
                y_reverse = False

            chart_attributes = {
                'title': plot_options[i]['chart_title'] if plot_options[i]['chart_title'] else None,
                'x_axis': {
                    'title': plot_options[i]['x_label'],
                    'crosses': 'max' if y_reverse else 'min',
                    'scaling': {
                        'min': x_min,
                        'max': x_max,
                        'orientation': 'maxMin' if x_reverse else 'minMax',
                        'logBase': 10 if plot_options[i]['x_log_scale'] else None
                    }
                },
                'y_axis': {
                    'title': plot_options[i]['y_label'],
                    'crosses': 'max' if x_reverse else 'min',
                    'scaling': {
                        'min': y_min,
                        'max': y_max,
                        'orientation': 'maxMin' if y_reverse else 'minMax',
                        'logBase': 10 if plot_options[i]['y_log_scale'] else None
                    }
                }
            }

            chart = ScatterChart()
            for key, attribute in chart_attributes.items():
                if not isinstance(attribute, dict):
                    setattr(chart, key, attribute)
                else:
                    for axis_attribute, value in attribute.items():
                        if not isinstance(value, dict):
                            setattr(getattr(chart, key), axis_attribute, value)
                        else:
                            for internal_attribute, internal_value in value.items():
                                setattr(
                                    getattr(getattr(chart, key), axis_attribute),
                                    internal_attribute, internal_value
                                )

            location = first_column
            for j in range(len(labels[i]['sample_names'])):
                for k in range(len(data_source.lengths[i][j])):
                    if plot_options[i][f'plot_{j}_{k}']:
                        series = Series(
                            Reference(
                                worksheet,
                                first_column + plot_options[i][f'y_plot_index_{j}_{k}'],
                                first_row + 2,
                                first_column + plot_options[i][f'y_plot_index_{j}_{k}'],
                                last_row
                            ),
                            xvalues=Reference(
                                worksheet,
                                first_column + plot_options[i][f'x_plot_index_{j}_{k}'],
                                first_row + 2,
                                first_column + plot_options[i][f'x_plot_index_{j}_{k}'],
                                last_row
                            )
                        )
                        series.title = SeriesLabel(
                            StrRef(f"'{sheet_name}'!{utils.excel_column_name(location)}{first_row}")
                        )
                        chart.append(series)
                location += sum(data_source.lengths[i][j])

            # default position is D8
            worksheet.add_chart(chart, f'{utils.excel_column_name(first_column + 3)}{first_row + 7}')


def _select_processing_options(data_sources):
    """
    Launches a window to select the processing options.

    Parameters
    ----------
    data_sources : list(DataSource) or tuple(DataSource)
        A container (list, tuple) of DataSource objects.

    Returns
    -------
    values : dict
        A dictionary containing the processing options.

    """

    options_layout = [
        [sg.Text('File Selection', relief='ridge', justification='center',
                 size=(40, 1))],
        [sg.Radio('Manually Select Files', 'options_radio', default=True,
                  key='manual_search')],
        [sg.Radio('Search Files Using Keywords', 'options_radio', key='keyword_search')],
        [sg.Radio('Use Previous Files', 'options_radio', key='use_last_search',
                  disabled=True)],
        [sg.Text('Select All Boxes That Apply', relief='ridge',
                 justification='center', size=(40, 1))],
        [sg.Check('Process Data', key='process_data', default=True,
                  enable_events=True)],
        [sg.Check('Fit Data', key='fit_data', enable_events=True)],
        [sg.Check('Save Results to Excel', key='save_fitting', pad=((40, 0), (1, 0)),
                  enable_events=True, disabled=True)],
        [sg.Check('Plot in Python', key='plot_python')],
        [sg.Check('Move File(s)', key='move_files', default=False)],
        [sg.Check('Save Excel File', key='save_excel',
                  default=True, enable_events=True),
         sg.Combo(('Create new file', 'Append to existing file'),
                  key='append_file', readonly=True,
                  default_value='Append to existing file', size=(19, 1))],
        [sg.Check('Plot Data in Excel', key='plot_data_excel',
                  pad=((40, 0), (1, 0)))],
        [sg.Check('Plot Fit Results in Excel', key='plot_fit_excel',
                  disabled=True, pad=((40, 0), (1, 0)))],
        [sg.Input('', key='file_name', visible=False),
         sg.Input('', key='display_name', disabled=True, size=(20, 1), pad=((40, 0), 5)),
         sg.Button('Save As', key='save_as')],
    ]

    data_sources_radios = [
        [sg.Radio(textwrap.fill(f'{source.name}', 30), 'radio', key=f'source_{source.name}',
                  enable_events=True)] for j, source in enumerate(data_sources)
    ]

    layout = [
        [sg.TabGroup([
            [sg.Tab('Data Sources', [
                [sg.Text('Select Data Source', relief='ridge',
                         justification='center', size=(40, 1))],
                [sg.Column(data_sources_radios, scrollable=True, vertical_scroll_only=True,
                           element_justification='left', key='DataSource_column')]
             ], key='tab1'),
             sg.Tab('Options', options_layout, key='tab2')]
        ], tab_background_color=sg.theme_background_color(), key='tab')],
        [sg.Button('Next', bind_return_key=True,
                   button_color=utils.PROCEED_COLOR)]
    ]

    window = sg.Window('Main Menu', layout, finalize=True, icon=utils._LOGO)
    window['DataSource_column'].expand(expand_x=True, expand_y=True)
    data_source = None
    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED:
            utils.safely_close_window(window)

        elif event.startswith('source'):
            for source in data_sources:
                if values[f'source_{source.name}']:
                    data_source = source
                    break

            if SAVE_FOLDER.joinpath(f'{_FILE_PREFIX}{data_source.name}.json').exists():
                window['use_last_search'].update(disabled=False)
            else:
                window['use_last_search'].update(value=False, disabled=True)
                if values['use_last_search']:
                    window['manual_search'].update(value=True)

        elif event == 'fit_data':
            if values['fit_data'] and values['save_excel']:
                window['save_fitting'].update(value=True, disabled=False)
                window['plot_fit_excel'].update(disabled=False)
            else:
                window['save_fitting'].update(value=False, disabled=True)
                window['plot_fit_excel'].update(value=False, disabled=True)

        elif event == 'save_fitting':
            if values['save_fitting']:
                window['plot_fit_excel'].update(disabled=False)
            else:
                window['plot_fit_excel'].update(value=False, disabled=True)

        elif event == 'save_excel':
            if values['save_excel']:
                window['display_name'].update(value=Path(values['file_name']).name)
                window['save_as'].update(disabled=False)
                window['append_file'].update(readonly=True)
                window['plot_data_excel'].update(disabled=False)

                if values['fit_data']:
                    window['save_fitting'].update(value=True, disabled=False)
                    window['plot_fit_excel'].update(disabled=False)
            else:
                window['display_name'].update(value='')
                window['save_as'].update(disabled=True)
                window['append_file'].update(
                    value='Append to existing file', disabled=True
                )
                window['plot_data_excel'].update(value=False, disabled=True)
                window['plot_fit_excel'].update(value=False, disabled=True)
                window['save_fitting'].update(value=False, disabled=True)

        elif event == 'save_as':
            file_name = sg.popup_get_file(
                '', save_as=True, default_path=values['display_name'], no_window=True,
                file_types=(("Excel Workbook (xlsx)", "*.xlsx"),), icon=utils._LOGO
            )
            if file_name:
                file_path = Path(file_name)
                if file_path.suffix.lower() != '.xlsx':
                    file_path = Path(file_path.parent, file_path.stem + '.xlsx')
                window['file_name'].update(value=str(file_path))
                window['display_name'].update(value=file_path.name)

        elif event == 'Next':
            if data_source is None:
                sg.popup('Please select a data source.\n', title='Error', icon=utils._LOGO)
            elif (not any(values[key] for key in ('process_data', 'fit_data', 'plot_python',
                    'move_files', 'save_excel'))):
                sg.popup('Please select a data processing option.\n',
                         title='Error', icon=utils._LOGO)
            elif values['save_excel'] and not values['file_name']:
                sg.popup('Please select a filename for the output Excel file.\n',
                         title='Error', icon=utils._LOGO)
            else:
                break

    window.close()
    del window

    values['append_file'] = values['append_file'] == 'Append to existing file'

    return values


def _create_column_labels_window(dataset, data_source, options, index,
                                 gui_inputs, location, last_index):
    """
    Creates the window to specify the sample and column labels.

    Parameters
    ----------
    dataset : list
        The list of lists of dataframes for one dataset.
    data_source : DataSource
        The DataSource object for the data.
    options : dict
        The dictionary that contains information about which
        processing steps will be conducted.
    index : int
        The index of the dataset within the total list of datasets.
    gui_inputs : dict
        A dictionary of values to overwrite the default gui values, used
        when displaying a previous window.
    location : tuple(int, int)
        The window location.
    last_index : bool
        If True, designates that it is the last index.

    Returns
    -------
    validations : dict
        A dictionary with the validations needed for the created window.
    sg.Window
        The created window to select the labels.

    """

    function_labels = data_source._create_function_labels() if options['process_data'] else [[], [], []]
    available_cols = data_source._create_data_labels() + function_labels[0]
    if (data_source.x_plot_index >= len(available_cols)
            or data_source.y_plot_index >= len(available_cols)):
        x_plot_index = 0
        y_plot_index = len(available_cols) - 1
    else:
        x_plot_index = data_source.x_plot_index
        y_plot_index = data_source.y_plot_index

    validations = {'user_inputs': []}
    default_inputs = {
        'sheet_name': f'Sheet {index + 1}',
        'x_min': '',
        'x_max': '',
        'y_min': '',
        'y_max': '',
        'x_label': available_cols[x_plot_index],
        'y_label': available_cols[y_plot_index],
        'x_log_scale': False,
        'y_log_scale': False,
        'chart_title': ''
    }

    total_labels = {f'Sample {i + 1}': {} for i in range(len(dataset))}
    column_count = 0
    for i, sample in enumerate(dataset):
        key = f'Sample {i + 1}'
        default_inputs.update({f'sample_name_{i}': ''})
        validations['user_inputs'].append([
            f'sample_name_{i}', f'sample name {i + 1}', utils.string_to_unicode, True, None
        ])

        for j, entry in enumerate(sample):
            subkey = f'Entry {j + 1}'
            total_labels[key][subkey] = list(itertools.chain(
                data_source._create_data_labels(len(entry.columns), options['process_data']),
                function_labels[0]
            ))

            entry_x_index = x_plot_index if x_plot_index < len(total_labels[key][subkey]) else 0
            entry_y_index = y_plot_index if y_plot_index < len(total_labels[key][subkey]) else len(total_labels[key][subkey]) - 1
            default_inputs.update({
                f'plot_{i}_{j}': True,
                f'x_plot_index_{i}_{j}': entry_x_index + column_count,
                f'y_plot_index_{i}_{j}': entry_y_index + column_count
            })

            for k, label in enumerate(total_labels[key][subkey]):
                if options['process_data'] and data_source.label_entries and len(sample) > 1 and label:
                    column_label = f'{label}, {j + 1}'
                else:
                    column_label = label

                default_inputs.update({f'column_name_{i}_{j}_{k}': column_label})
                validations['user_inputs'].append([
                    f'column_name_{i}_{j}_{k}', f'column name {column_count}',
                    utils.string_to_unicode, True, None
                ])
                column_count += 1

        if function_labels[1]:
            subkey = 'Sample Summary'
            total_labels[key][subkey] = function_labels[1]
            entry_x_index = x_plot_index if x_plot_index < len(function_labels[1]) else 0
            entry_y_index = y_plot_index if y_plot_index < len(function_labels[1]) else len(function_labels[1]) - 1
            default_inputs.update({
                f'plot_{i}_{j + 1}': False,
                f'x_plot_index_{i}_{j + 1}': entry_x_index + column_count,
                f'y_plot_index_{i}_{j + 1}': entry_y_index + column_count
            })

            for k, label in enumerate(function_labels[1]):
                default_inputs.update({f'column_name_{i}_{j + 1}_{k}': label})
                validations['user_inputs'].append([
                    f'column_name_{i}_{j + 1}_{k}', f'column name {column_count}',
                    utils.string_to_unicode, True, None
                ])
                column_count += 1

    if function_labels[2]:
        default_inputs.update({'summary_name': 'Summary'})
        validations['user_inputs'].append([
            'summary_name', 'summary name', utils.string_to_unicode, True, None
        ])
        total_labels['Dataset Summary'] = {}
        total_labels['Dataset Summary']['Entry 1'] = function_labels[2]
        entry_x_index = x_plot_index if x_plot_index < len(function_labels[2]) else 0
        entry_y_index = y_plot_index if y_plot_index < len(function_labels[2]) else len(function_labels[2]) - 1
        default_inputs.update({
            f'plot_{i + 1}_0': False,
            f'x_plot_index_{i + 1}_0': entry_x_index + column_count,
            f'y_plot_index_{i + 1}_0': entry_y_index + column_count
        })
        for k, label in enumerate(function_labels[2]):
            default_inputs.update({f'column_name_{i + 1}_0_{k}': label})
            validations['user_inputs'].append([
                f'column_name_{i + 1}_0_{k}', f'column name {column_count}',
                utils.string_to_unicode, True, None
            ])
            column_count += 1

    # overwrites the defaults with any previous inputs
    default_inputs.update(gui_inputs)

    if options['save_excel']:
        header = 'Sheet Name: '
        validations['user_inputs'].extend([
            ['sheet_name', 'sheet name', utils.string_to_unicode, False, None],
            ['sheet_name', 'sheet name', utils.validate_sheet_name, False, None]
        ])
    else:
        header = f'Dataset {index + 1}'

    input_width = 25
    labels_layout = [
        [sg.Text(header, visible=options['save_excel']),
         sg.Input(default_inputs['sheet_name'], key='sheet_name',
                  size=(input_width, 1), visible=options['save_excel'])]
    ]
    headers = []
    for i, sample_name in enumerate(total_labels.keys()):
        key = f'sample_name_{i}' if sample_name != 'Dataset Summary' else 'summary_name'
        headers.append([
            sg.Text(f'  {sample_name}'),
            sg.Input(default_inputs[key], size=(input_width, 1), key=key),
            sg.Text('  ')
        ])

    labels_layout.append([
        sg.Frame('Sample Names', [[
            sg.Column(
                headers,
                scrollable=True,
                vertical_scroll_only=True,
                expand_x=True,
                element_justification='center',
                size=(None, 150)
            )
        ]], element_justification='center', title_location=sg.TITLE_LOCATION_TOP)
    ])

    column_labels = []
    column_count = 0
    for i, (sample_name, sample_values) in enumerate(total_labels.items()):
        column_labels.append(
            [sg.Column([
                [sg.HorizontalSeparator()],
                [sg.Text(sample_name)],
                [sg.HorizontalSeparator()]
            ], expand_x=True, element_justification='center')]
        )
        for j, (entry_label, label_list) in enumerate(sample_values.items()):
            column_labels.append(
                [sg.Column([[sg.Text(entry_label)]],
                           expand_x=True, element_justification='center')]
            )
            for k, label in enumerate(label_list):
                column_labels.append([
                    sg.Text(f'  Column {column_count}'),
                    sg.Input(default_inputs[f'column_name_{i}_{j}_{k}'],
                             size=(input_width, 1), key=f'column_name_{i}_{j}_{k}'),
                    sg.Text('  ')
                ])
                column_count += 1

    labels_layout.append([
        sg.Frame('Column Labels', [[
            sg.Column(
                column_labels,
                scrollable=True,
                vertical_scroll_only=True,
                expand_x=True,
                element_justification='center',
                size=(None, 150)
            )
        ]], element_justification='center', title_location=sg.TITLE_LOCATION_TOP)
    ])

    if not options['plot_data_excel']:
        main_section = [sg.Column(labels_layout)]
    else:
        validations['integers'] = []
        validations['user_inputs'].extend([
            ['x_min', 'x min', float, True, None],
            ['x_max', 'x max', float, True, None],
            ['y_min', 'y min', float, True, None],
            ['y_max', 'y max', float, True, None],
            ['x_label', 'x axis label', utils.string_to_unicode, False, None],
            ['y_label', 'y axis label', utils.string_to_unicode, False, None],
            ['chart_title', 'chart title', utils.string_to_unicode, True, None]
        ])

        total_indices = list(range(column_count))
        plot_indices = []
        for i, (sample_name, sample_values) in enumerate(total_labels.items()):
            for j, entry_label in enumerate(sample_values.keys()):
                validations['integers'].extend([
                    [f'x_plot_index_{i}_{j}', 'x plot index'],
                    [f'y_plot_index_{i}_{j}', 'y plot index']
                ])

                plot_indices.extend([
                    [sg.Check(f'Plot {sample_name}, {entry_label}', default_inputs[f'plot_{i}_{j}'],
                              key=f'plot_{i}_{j}')],
                    [sg.Text('    X column'),
                     sg.Combo(total_indices, total_indices[default_inputs[f'x_plot_index_{i}_{j}']],
                              readonly=True, key=f'x_plot_index_{i}_{j}', size=(5, 1)),
                     sg.Text(' Y column'),
                     sg.Combo(total_indices, total_indices[default_inputs[f'y_plot_index_{i}_{j}']],
                              readonly=True, key=f'y_plot_index_{i}_{j}', size=(5, 1))]
                ])

        plot_layout = [
            [sg.Text('Chart title:'),
             sg.Input(default_inputs['chart_title'], key='chart_title',
                      size=(input_width, 1))],
            [sg.Text('X axis label:'),
             sg.Input(default_inputs['x_label'], key='x_label', size=(input_width, 1))],
            [sg.Text('Y axis label:'),
             sg.Input(default_inputs['y_label'], key='y_label', size=(input_width, 1))],
            [sg.Text(('Min and max values to show on the plot\n'
                      "(leave blank to use Excel's default):"))],
            [sg.Text('    X min:', size=(8, 1)),
             sg.Input(default_inputs['x_min'], key='x_min', size=(5, 1)),
             sg.Text('    X max:', size=(8, 1)),
             sg.Input(default_inputs['x_max'], key='x_max', size=(5, 1))],
            [sg.Text('    Y min:', size=(8, 1)),
             sg.Input(default_inputs['y_min'], key='y_min', size=(5, 1)),
             sg.Text('    Y max:', size=(8, 1)),
             sg.Input(default_inputs['y_max'], key='y_max', size=(5, 1))],
            [sg.Text('Use logorithmic scale?')],
            [sg.Check('X axis', default_inputs['x_log_scale'],
                      key='x_log_scale', pad=((20, 5), 5)),
             sg.Check('Y axis', default_inputs['y_log_scale'], key='y_log_scale')],
            [sg.Frame('', [[
                sg.Column(plot_indices, scrollable=True, vertical_scroll_only=True, size=(None, 150))
            ]])]
        ]

        main_section = [
            sg.TabGroup([
                [sg.Tab('Labels', labels_layout, key='tab1'),
                 sg.Tab('Excel Plot', plot_layout, key='tab2')]
            ], key='tab', tab_background_color=sg.theme_background_color())
        ]

    layout = [
        [sg.Menu([['&Help', ['&Unicode Help']]], key='-menu-',
                 background_color=sg.theme_background_color())],
        main_section,
        [sg.Text('')],
        [sg.Button('Back', disabled=index == 0),
         sg.Button('Finish' if last_index else 'Next', bind_return_key=True,
                   button_color=utils.PROCEED_COLOR)]
    ]

    return validations, sg.Window(f'Dataset {index + 1} Options', layout, location=location, resizable=True, icon=utils._LOGO)


def _select_column_labels(dataframes, data_source, processing_options):
    """
    Handles the event loop for the window to select the sample and column labels.

    Parameters
    ----------
    dataframes : list
        A list of lists of lists of pd.DataFrame objects, containing the all
        the data to process.
    data_source : DataSource
        The DataSource object for the data.
    processing_options : dict
        The dictionary that contains information about which
        processing steps will be conducted.

    Returns
    -------
    labels : list(dict)
        A list of dictionaries containing all of the sample and column
        labels, as well as the Excel plot options, if plotting in Excel.
        Each entry in the list corresponds to one dataset.
    plot_options : list(dict)
        A list of dictionaries with values used to create the Excel plot
        if plot_excel is True.

    """

    non_transferred_keys = ['sheet_name', 'plot_', 'x_plot_index_', 'y_plot_index_']
    if len(set(len(df.columns) for dataset in dataframes for sample in dataset for df in sample)) > 1:
        # don't transfer column names if column lengths are not all the same
        non_transferred_keys.append('column_name')
    non_transferred_keys = tuple(non_transferred_keys) # so it works with str.startswith

    label_values = [{} for _ in dataframes]
    location = (None, None)
    for i, dataset in enumerate(dataframes):
        j = i

        validations, window = _create_column_labels_window(
            dataset, data_source, processing_options, i, label_values[i],
            location, i == len(dataframes) - 1
        )

        while True:
            event, values = window.read()

            if event == sg.WIN_CLOSED:
                utils.safely_close_window(window)

            elif event == 'Unicode Help':
                sg.popup(
                    ('"\\u00B2": \u00B2 \n"\\u03B8": \u03B8 \n"'
                     '\\u00B0": \u00B0\n"\\u03bc": \u03bc\n"\\u03bb": \u03bb\n'
                     '\nFor example, Acceleration'
                     ' (m/s\\u00B2) creates Acceleration (m/s\u00B2).\n'),
                    title='Example Unicode', modal=False, icon=utils._LOGO
                )

            elif (event in ('Back', 'Next', 'Finish')
                  and utils.validate_inputs(values, **validations)):

                label_values[j] = values
                location = window.current_location()
                window.close()

                if event == 'Back':
                    j -= 1
                else:
                    j += 1

                if j <= i:
                    validations, window = _create_column_labels_window(
                        dataframes[j], data_source, processing_options, j,
                        label_values[j], location, j == len(dataframes) - 1
                    )
                else:
                    if i < len(dataframes) - 1:
                        transfer_keys = set(
                            key for key in values.keys() if not key.startswith(non_transferred_keys)
                        )
                        label_values[i + 1].update({key: val for key, val in values.items() if key in transfer_keys})
                    break

        window.close()
        window = None

    plot_options = [{} for _ in label_values]
    if processing_options['plot_data_excel']:
        for i, values in enumerate(label_values):
            plot_options[i].update({
                'x_label': values['x_label'],
                'y_label': values['y_label'],
                'chart_title' : values['chart_title'],
                'x_min': values['x_min'] if values['x_min'] != '' else None,
                'x_max': values['x_max'] if values['x_max'] != '' else None,
                'y_min': values['y_min'] if values['y_min'] != '' else None,
                'y_max': values['y_max'] if values['y_max'] != '' else None,
                'x_log_scale': values['x_log_scale'],
                'y_log_scale': values['y_log_scale']
            })
            plot_options[i].update(
                {key: value for key, value in values.items()
                 if key.startswith(('plot_', 'x_plot_index', 'y_plot_index'))}
            )

    labels, plot_options = _collect_column_labels(label_values, plot_options, data_source, processing_options)

    return labels, plot_options


def _collect_column_labels(label_values, plot_options, data_source, options):
    """
    Collects all labels and condenses them into a single list of labels per dataset.

    Also adds in blank labels for spacer columns between entries and samples and
    adjusts the indices for plotting accordingly.

    Parameters
    ----------
    label_values : list(dict)
        A list of dictionaries. Each dictionary contains all of the
        sample names and column labels for a dataset.
    data_source : DataSource
        The DataSource object for the data.
    options : dict
        The dictionary that contains information about which
        processing steps will be conducted.

    Returns
    -------
    labels : list(dict)
        A list of dictionaries for each dataset. Each internal dictionary
        contains sheet_name, sample_names and columns_names for writing
        each dataset to Excel, and dataframe_names for the dataframe columns.
    plot_options : list(dict)
        A list of dictionaries with values used to create the Excel plot
        if plot_excel is True.

    """

    labels = [{} for _ in label_values]
    for num, label_dict in enumerate(label_values):
        labels[num]['sheet_name'] = label_dict.get('sheet_name', '')

        sample_keys = [key for key in label_dict.keys() if key.startswith('sample_name')]
        labels[num]['sample_names'] = [label_dict[f'sample_name_{i}'] for i in range(len(sample_keys))]
        if 'summary_name' in label_dict:
            labels[num]['sample_names'].append(label_dict['summary_name'])

        plot_indices = {key: value for key, value in plot_options[num].items() if 'plot_index_' in key}
        labels[num]['column_names'] = []
        labels[num]['dataframe_names'] = []
        column_index = 0
        for i in range(len(labels[num]['sample_names'])):
            entries = 1 + max([
                int(key.split('_')[-2]) for key in label_dict.keys()
                if key.startswith(f'column_name_{i}')
            ])
            for j in range(entries):
                columns = len([key for key in label_dict.keys() if key.startswith(f'column_name_{i}_{j}_')])
                labels[num]['column_names'].extend([
                    label_dict[f'column_name_{i}_{j}_{k}'] for k in range(columns)
                ])
                labels[num]['dataframe_names'].extend([
                    label_dict[f'column_name_{i}_{j}_{k}'] for k in range(columns)
                ])
                column_index += columns

                if options['process_data'] and j != entries - 1:
                    labels[num]['column_names'].extend([
                        '' for _ in range(data_source.entry_separation)
                    ])
                    if options['plot_data_excel']:
                        for key, value in plot_indices.items():
                            if value >= column_index:
                                plot_indices[key] += data_source.entry_separation
                    column_index += data_source.entry_separation

            if options['process_data']:
                labels[num]['column_names'].extend(['' for _ in range(data_source.sample_separation)])
                if options['plot_data_excel']:
                    for key, value in plot_indices.items():
                        if value >= column_index:
                            plot_indices[key] += data_source.sample_separation
                    column_index += data_source.sample_separation

        plot_options[num].update(plot_indices)

    return labels, plot_options


def _fit_data(datasets, data_source, labels,
              excel_writer, options, rc_params=None):
    """
    Handles fitting the data and any exceptions that occur during fitting.

    Parameters
    ----------
    dataframes : list
        A nested list of lists of lists of dataframes.
    data_source : DataSource
        The selected DataSource.
    labels : list(dict)
        A list of dictionaries containing the sample names and column
        labels for each dataset.
    excel_writer : pd.ExcelWriter
        The pandas ExcelWriter object that contains all of the
        information about the Excel file being created.
    options : dict
        A dictionary containing the relevent keys 'save_fitting' and
        'plot_fit_excel' which determine whether the fit results
        will be saved to Excel and whether the results will be plotted,
        respectively.
    rc_params : dict, optional
        A dictionary of changes to matplotlib's rcparams. If None, will
        use data_source.figure_rcparams.

    Returns
    -------
    results : list(list(list(lmfit.models.ModelResult or None)))
        A nested list of lists of lists of lmfit.ModelResults, one for each
        entry in each sample in each dataset in datasets. If fitting was not
        done for the entry, the value will be None.

    Raises
    ------
    utils.WindowCloseError
        Raised if fitting was ended early by the user.

    """

    from .fitting import launch_fitting_gui


    if rc_params is not None:
        mpl_changes = rc_params.copy()
    else:
        mpl_changes = data_source.figure_rcparams.copy()

    results = [[[] for sample in dataset] for dataset in datasets]

    # Allows exiting from the peak fitting GUI early, if desired or because of
    # an exception, while still continuing with the program.
    try:
        default_inputs = {
            'x_fit_index': data_source.x_plot_index,
            'y_fit_index': data_source.y_plot_index
        }

        for i, dataset in enumerate(datasets):
            default_inputs.update({
                'x_label': labels[i]['column_names'][data_source.x_plot_index],
                'y_label': labels[i]['column_names'][data_source.y_plot_index]
            })
            sample_names = labels[i]['sample_names']
            for j, sample in enumerate(dataset):
                for k, entry in enumerate(sample):
                    if len(sample) > 1:
                        name = f'{sample_names[j]}_{k + 1}_fit'
                    else:
                        name = sample_names[j]

                    default_inputs.update({'sample_name': name})

                    fit_output, default_inputs, proceed = launch_fitting_gui(
                        entry, default_inputs, excel_writer,
                        options['save_fitting'], options['plot_fit_excel'],
                        mpl_changes, False, data_source.excel_styles
                    )

                    results[i][j].extend(fit_output)

                    if not proceed:
                        raise utils.WindowCloseError

    except (utils.WindowCloseError, KeyboardInterrupt):
        print('\nPeak fitting manually ended early.\nMoving on with program.')

    except Exception:
        print('\nException occured during peak fitting:\n')
        print(traceback.format_exc())
        print('Moving on with program.')

    return results


def _plot_data(datasets, data_source):
    """
    Handles plotting and any exceptions that occur during plotting.

    Parameters
    ----------
    datasets : list
        A nested list of lists of lists of dataframes.
    data_source : DataSource
        The DataSource object whose figure_rcparams attribute will be used
        to set matplotlib's rcParams.

    Returns
    -------
    list
        A nested list of lists, with one entry per dataset in datasets.
        Each entry contains the matplotlib Figure, and a dictionary
        containing the Axes. If plotting was exited before plotting all
        datasets in dataframes, then [None, None] will be the entry instead.

    """

    from .plotting import launch_plotting_gui


    plot_datasets = []
    for dataset in datasets: # Flattens the dataset to a single list per dataset
        plot_datasets.append(list(itertools.chain.from_iterable(dataset)))

    return launch_plotting_gui(plot_datasets, data_source.figure_rcparams)


def _move_files(files):
    """
    Launches a window to select the new folder destinations for the files.

    Parameters
    ----------
    files : list
        A nested list of lists of lists of strings corresponding
        to file paths.

    """

    text_layout = [[sg.Text(f'Dataset {i + 1}')] for i in range(len(files))]
    files_layout = [
        [sg.Input('', key=f'folder_{i}', enable_events=True,
                  disabled=True),
         sg.FolderBrowse(target=f'folder_{i}', key=f'button_{i}')]
        for i in range(len(files))
    ]
    tot_layout = [i for j in zip(text_layout, files_layout) for i in j]

    if len(files) > 2:
        scrollable = True
        size = (600, 200)
    else:
        scrollable = False
        size = (None, None)

    layout = [
        [sg.Text('Choose the folder(s) to move files to:', size=(30, 1))],
        [sg.Frame('', [[sg.Column(tot_layout, scrollable=scrollable,
                                  vertical_scroll_only=True, size=size)]])],
        [sg.Button('Submit', bind_return_key=True,
                   button_color=utils.PROCEED_COLOR),
         sg.Check('All Same Folder', key='same_folder',
                  enable_events=True, disabled=len(files) == 1)]
    ]

    try:
        window = sg.Window('Move Files', layout, icon=utils._LOGO)
        while True:
            event, values = window.read()

            if event == sg.WIN_CLOSED:
                utils.safely_close_window(window)

            elif event.startswith('folder_') and values['same_folder']:
                for i in range(1, len(files)):
                    window[f'folder_{i}'].update(value=values['folder_0'])

            elif event == 'same_folder':
                if values['same_folder']:
                    for i in range(1, len(files)):
                        window[f'folder_{i}'].update(value=values['folder_0'])
                        window[f'button_{i}'].update(disabled=True)
                else:
                    for i in range(1, len(files)):
                        window[f'button_{i}'].update(disabled=False)

            elif event == 'Submit':
                if any(not values[key] for key in values if key.startswith('folder_')):
                    sg.popup('Please enter folders for all datasets',
                             title='Error', icon=utils._LOGO)
                else:
                    break

        window.close()
        del window

    except (utils.WindowCloseError, KeyboardInterrupt):
        print('\nMoving files manually ended early.\nMoving on with program.')

    else:
        try:
            folders = [values[f'folder_{i}'] for i in range(len(files))]
            for i, file_list in enumerate(files):
                # Will automatically rename files if there is already a file with
                # the same name in the destination folder.
                file_mover(file_list, new_folder=folders[i], skip_same_files=False)
        except Exception:
            print('\nException occured during moving files:\n')
            print(traceback.format_exc())
            print('Moving on with program.')


def launch_main_gui(data_sources, fitting_mpl_params=None):
    """
    Goes through all steps to find files, process/fit/plot the imported data, and save to Excel.

    Parameters
    ----------
    data_sources : list(DataSource) or tuple(DataSource) or DataSource
        A list or tuple of mcetl.DataSource objects, or a single DataSource.
    fitting_mpl_params : dict, optional
        A dictionary of changes for Matplotlib's rcParams to use
        during fitting. If None, will use the selected DataSource's
        figure_rcparams attribute.

    Returns
    -------
    output : dict
        A dictionary containing the following keys and values:

            'dataframes': list or None
                A list of lists of dataframes, with each dataframe containing the
                data imported from a raw data file; will be None if the function
                fails before importing data, or if the only processing step taken
                was moving files.
            'fit_results': list or None
                A nested list of lists of lmfit.ModelResult objects, with each
                ModelResult pertaining to the fitting of a data entry, each list of
                ModelResults containing all of the fits for a single sample,
                and east list of lists pertaining to the data within one dataset.
                Will be None if fitting is not done, or only partially filled
                if the fitting process ends early.
            'plot_results': list or None
                A list of lists, with one entry per dataset. Each interior
                list is composed of a matplotlib.Figure object and a
                dictionary of matplotlib.Axes objects. Will be None if
                plotting is not done, or only partially filled if the plotting
                process ends early.
            'writer': pd.ExcelWriter or None
                The pandas ExcelWriter used to create the output Excel file; will
                be None if the output results were not saved to Excel.

    Notes
    -----
    The entire function is wrapped in a try-except block. If the user exits the
    program early by exiting out of a GUI, a custom WindowCloseError exception is
    thrown, which is just passed, allowing the program is close without error.
    If other exceptions occur, their traceback is printed.

    """

    output = {
        'dataframes': None,
        'fit_results': None,
        'plot_results': None,
        'writer': None
    }

    if not isinstance(data_sources, (list, tuple)):
        data_sources = [data_sources]
    if any(not isinstance(data_source, DataSource) for data_source in data_sources):
        raise TypeError("Only mcetl.DataSource objects can be used in mcetl's main gui.")

    try:
        processing_options = _select_processing_options(data_sources)

        # Specifying the selected data source
        for source in data_sources:
            if processing_options[f'source_{source.name}']:
                data_source = source
                break
        # Removes unique variables that are only used in preprocessing
        if not processing_options['process_data']:
            data_source._remove_unneeded_variables()

        # Selection of data files
        if processing_options['use_last_search']:
            with SAVE_FOLDER.joinpath(f'{_FILE_PREFIX}{data_source.name}.json').open('r') as fp:
                files = json.load(fp)
        else:
            if processing_options['keyword_search']:
                files = file_finder(
                    file_type=data_source.file_type, num_files=data_source.num_files
                )
            else:
                files = manual_file_finder(data_source.file_type)

            # Saves the file paths to a json file so they can be used again to bypass the search.
            try:
                SAVE_FOLDER.mkdir(exist_ok=True)
                with SAVE_FOLDER.joinpath(f'{_FILE_PREFIX}{data_source.name}.json').open('w') as fp:
                    json.dump(files, fp, indent=2)
            except PermissionError:
                # do not create the folder and/or files if cannot access
                warnings.warn((
                    f'Write access is denied in {str(SAVE_FOLDER)}, so '
                    f'{_FILE_PREFIX}{data_source.name}.json was not written.'
                ))

        # Imports the raw data from the files and specifies column names
        if any((processing_options['process_data'],
                processing_options['save_excel'],
                processing_options['fit_data'],
                processing_options['plot_python'])):

            output['dataframes'] = [[[] for sample in dataset] for dataset in files]
            references = [[[] for sample in dataset] for dataset in files]
            import_values = {}
            for i, dataset in enumerate(files):
                for j, sample in enumerate(dataset):
                    for entry in sample:
                        if (not import_values.get('same_values', False)
                                or Path(entry).suffix.lower() in utils._get_excel_engines()):
                            import_values = utils.select_file_gui(
                                data_source, entry, import_values,
                                processing_options['process_data']
                            )

                        added_dataframes = utils.raw_data_import(import_values, entry, False)
                        output['dataframes'][i][j].extend(added_dataframes)
                        import_vals = {}
                        for var in data_source.unique_variables:
                            # use .get() since keys will not exist if not processing
                            import_vals[var] = import_values.get(f'index_{var}')
                        references[i][j].extend([import_vals] * len(added_dataframes))

            if processing_options['process_data']:
                # Perform preprocessing functions before assigning column labels
                # since columns could be added/removed
                output['dataframes'], references = data_source._do_preprocessing(
                    output['dataframes'], references
                )

            labels, plot_options = _select_column_labels(
                output['dataframes'], data_source, processing_options
            )

        if processing_options['save_excel'] or processing_options['process_data']:
            if processing_options['process_data']:
                # Assign reference indices for all relevant columns
                data_source._set_references(output['dataframes'], references)

            # Merge dataframes for each dataset
            merged_dataframes = data_source.merge_datasets(output['dataframes'])
            output['dataframes'] = None # Frees up memory

            if processing_options['save_excel'] and processing_options['process_data']:
                merged_dataframes = data_source._do_excel_functions(merged_dataframes)

            if processing_options['save_excel']:
                # Create the writer handler and read the Excel file if appending.
                writer_handler = ExcelWriterHandler(
                    processing_options['file_name'], not processing_options['append_file'],
                    data_source.excel_styles
                )
                output['writer'] = writer_handler.writer

                _write_to_excel(
                    merged_dataframes, data_source, labels, writer_handler,
                    processing_options['plot_data_excel'], plot_options
                )

            if processing_options['process_data']:
                merged_dataframes = data_source._do_python_functions(merged_dataframes)

            # Split data back into individual dataframes
            output['dataframes'] = data_source.split_into_entries(merged_dataframes)
            del merged_dataframes

        # Assign column headers for all dataframes
        if any((processing_options['process_data'],
                processing_options['save_excel'],
                processing_options['fit_data'],
                processing_options['plot_python'])):

            for i, dataset in enumerate(output['dataframes']):
                column_names = iter(labels[i]['dataframe_names'])
                for sample in dataset:
                    for entry in sample:
                        entry.columns = [next(column_names) for _ in range(len(entry.columns))]

        # Handles peak fitting
        if processing_options['fit_data']:
            output['fit_results'] = _fit_data(
                output['dataframes'], data_source, labels, output['writer'],
                processing_options, fitting_mpl_params
            )

        # Handles saving the Excel file
        if processing_options['save_excel']:
            writer_handler.save_excel_file()

        # Handles moving files
        if processing_options['move_files']:
            _move_files(files)

        # Handles plotting in python
        if processing_options['plot_python']:
            output['plot_results'] = _plot_data(output['dataframes'], data_source)

    except (utils.WindowCloseError, KeyboardInterrupt):
        pass
    except Exception:
        print(traceback.format_exc())

    return output
