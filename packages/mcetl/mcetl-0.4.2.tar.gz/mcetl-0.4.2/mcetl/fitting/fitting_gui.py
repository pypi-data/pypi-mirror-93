# -*- coding: utf-8 -*-
"""Provides a GUI to fit data using lmfit Models and save the results to Excel.

@author: Donald Erb
Created on May 24, 2020

Notes
-----
openpyxl is imported within fit_to_excel to reduce the import time of the module,
and is only imported if saving fit results to Excel.

"""


import itertools
from pathlib import Path
import traceback

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import PySimpleGUI as sg

from . import fitting_utils as f_utils
from . import peak_fitting
from .. import plot_utils, utils
from ..excel_writer import ExcelWriterHandler
# openpyxl is imported within fit_to_excel


class SimpleEmbeddedFigure(plot_utils.EmbeddedFigure):
    """
    A window containing just an embedded figure and a close button.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The dataframe that contains the x and y data.
    gui_values : dict
        A dictionary of values needed for plotting.

    """

    def __init__(self, dataframe, gui_values):

        x_data = utils.series_to_numpy(dataframe.iloc[:, gui_values['x_fit_index']])
        y_data = utils.series_to_numpy(dataframe.iloc[:, gui_values['y_fit_index']])
        super().__init__(x_data, y_data, enable_events=False)

        x_min = max(gui_values['x_min'], np.nanmin(x_data))
        x_max = min(gui_values['x_max'], np.nanmax(x_data))
        bkg_min = max(gui_values['bkg_x_min'], x_min)
        bkg_max = min(gui_values['bkg_x_max'], x_max)

        x_mid = (x_max + x_min) / 2
        bkg_mid = (bkg_max + bkg_min) / 2
        additional_peaks = np.array(gui_values['peak_list'])
        additional_peaks = additional_peaks[(additional_peaks > x_min)
                                            & (additional_peaks < x_max)]

        desired_dpi = 150
        dpi = plot_utils.determine_dpi(
            fig_width=self.canvas_size[0], fig_height=self.canvas_size[1],
            dpi=desired_dpi, canvas_size=self.canvas_size
        )
        self.figure, self.axis = plt.subplots(
            num='Fitting', tight_layout=True,
            figsize=np.array(self.canvas_size) / desired_dpi, dpi=dpi
        )
        self.axis.plot(x_data, y_data)
        ax_y = self.axis.get_ylim()
        self.axis.set_ylim(plot_utils.scale_axis(ax_y, 0.15, 0.15))

        peaks = _find_peaks(dataframe, gui_values)
        other_peaks = False
        for peak in peaks:
            if peak not in additional_peaks:
                other_peaks = True
                found_peaks = self.axis.vlines(
                    peak, *plot_utils.scale_axis(ax_y, 0.01, 0.03),
                    color='green', linestyle='-.', lw=2
                )
        for peak in additional_peaks:
            user_peaks = self.axis.vlines(
                peak, *plot_utils.scale_axis(ax_y, 0.01, 0.03),
                color='blue', linestyle=':', lw=2
                )
        self.axis.annotate(
            '', (x_max, plot_utils.scale_axis(ax_y, None, 0.03)[1]),
            (x_mid, plot_utils.scale_axis(ax_y, None, 0.03)[1]),
            arrowprops=dict(width=1.2, headwidth=5, headlength=5, color='black'),
            annotation_clip=False,
        )
        self.axis.annotate(
            '', (x_min, plot_utils.scale_axis(ax_y, None, 0.03)[1]),
            (x_mid, plot_utils.scale_axis(ax_y, None, 0.03)[1]),
            arrowprops=dict(width=1.2, headwidth=5, headlength=5, color='black'),
            annotation_clip=False,
        )
        self.axis.annotate(
            'Fitting range', (x_mid, plot_utils.scale_axis(ax_y, None, 0.063)[1]),
            ha='center'
        )
        self.axis.vlines(
            x_min, *plot_utils.scale_axis(ax_y, 0.01, 0.03),
            color='black', linestyle='-', lw=2
        )
        self.axis.vlines(
            x_max, *plot_utils.scale_axis(ax_y, 0.01, 0.03),
            color='black', linestyle='-', lw=2
        )

        if gui_values['subtract_bkg']:
            self.axis.annotate(
                '', (bkg_max, plot_utils.scale_axis(ax_y, 0.01, None)[0]),
                (bkg_mid, plot_utils.scale_axis(ax_y, 0.01, None)[0]),
                arrowprops=dict(width=1.2, headwidth=5, headlength=5, color='red'),
                annotation_clip=False,
            )
            self.axis.annotate(
                '', (bkg_min, plot_utils.scale_axis(ax_y, 0.01, None)[0]),
                (bkg_mid, plot_utils.scale_axis(ax_y, 0.01, None)[0]),
                arrowprops=dict(width=1.2, headwidth=5, headlength=5, color='red'),
                annotation_clip=False
            )
            self.axis.annotate(
                'Background range',
                (bkg_mid, plot_utils.scale_axis(ax_y, 0.085, None)[0]),
                color='red', ha='center'
            )
            self.axis.vlines(
                bkg_min, *plot_utils.scale_axis(ax_y, 0.01, 0.03),
                color='red', linestyle='--', lw=2
            )
            self.axis.vlines(
                bkg_max, *plot_utils.scale_axis(ax_y, 0.01, 0.03),
                color='red', linestyle='--', lw=2
            )

        peak_list = []
        if additional_peaks.size > 0 and other_peaks:
            peak_list = [found_peaks, user_peaks]
            label_list = ['Found peaks', 'User input peaks']
        elif additional_peaks.size > 0:
            peak_list = [user_peaks]
            label_list = ['User input peaks']
        elif peaks:
            peak_list = [found_peaks]
            label_list = ['Found peaks']

        if peak_list:
            self.axis.legend(
                peak_list, label_list, frameon=False, ncol=2,
                bbox_to_anchor=(0.0, 1.01, 1, 1.01), loc='lower left',
                borderaxespad=0, mode='expand'
            )

        self._create_window(window_title='Preview', button_text='Back')
        self._place_figure_on_canvas()


class ResultsPlot(plot_utils.EmbeddedFigure):
    """
    Shows the results of a fit to allow user to decide if fit was acceptable.

    Parameters
    ----------
    fit_result : lmfit.ModelResult
        The fit result to display.

    """

    def __init__(self, fit_result):
        x = fit_result.userkws['x']
        y = fit_result.data
        super().__init__(x, y, enable_events=False)
        self.canvas_size = (self.canvas_size[0] - 50, self.canvas_size[1] - 50)

        desired_dpi = 150
        dpi = plot_utils.determine_dpi(
            fig_width=self.canvas_size[0], fig_height=self.canvas_size[1],
            dpi=desired_dpi, canvas_size=self.canvas_size
        )
        self.figure, (self.axis, residual_ax) = plt.subplots(
            2, sharex=True, num='Fit Results', tight_layout=True,
            figsize=np.array(self.canvas_size) / desired_dpi, dpi=dpi,
            gridspec_kw={'height_ratios':[5, 1], 'hspace': 0}
        )

        self.axis.plot(x, y, 'o', color='dodgerblue', label='data')
        self.axis.plot([np.nan], [np.nan], 'ro', ms=1.5, label='residuals')
        self.axis.plot(self.x, fit_result.best_fit, 'k--', label='best fit', zorder=3)
        # don't plot fit_result.residual because it equals inf when there is a bad fit
        residual_ax.plot(x, y - fit_result.best_fit, 'ro', ms=1.5)
        residual_ax.axhline(0, color='k', linestyle='-', lw=1)
        residual_ax.set_ylim(plot_utils.scale_axis(residual_ax.get_ylim(), 0.05, 0.05))

        individual_models = fit_result.eval_components(x=x)
        if 'background_' in individual_models:
            background = individual_models.pop('background_')
            if isinstance(background, (int, float)) or len(background) == 1:
                background = np.full(self.x.shape[0], background)
        else:
            background = 0
        # Creates a color cycle to override matplotlib's to prevent color clashing
        self.axis.set_prop_cycle(color=['#ff7f0e', '#2ca02c', '#d62728', '#8c564b',
                                        '#e377c2', '#bcbd22', '#17becf'])
        for label, values in individual_models.items():
            if isinstance(values, (int, float)) or len(values) == 1:
                model_values = np.full(self.x.shape[0], values)
            else:
                model_values = values
            _plot_model_component(self.axis, self.x, model_values + background, label)

        if isinstance(background, np.ndarray) or background != 0:
            try:
                self.axis.plot(self.x, background, 'k-', label='background')
            except:
                self.background = None
            else:
                self.background = background
        else:
            self.background = None

        self.axis.legend(ncol=max(1, len(self.axis.lines) // 4))
        self._create_window(fit_result)
        self._place_figure_on_canvas()


    def _create_window(self, fit_result):
        """Creates a GUI with the figure canvas and two buttons."""

        self.toolbar_canvas = sg.Canvas(key='controls_canvas', pad=(0, (0, 10)),
                                        size=(self.canvas_size[0], 50))
        self.canvas = sg.Canvas(key='fig_canvas', size=self.canvas_size, pad=(0, 0))
        plot_tab = sg.Tab(
            'Plot', [
                [self.canvas],
                [self.toolbar_canvas],
                [sg.Check('Preview without background', key='subtract_bkg',
                          enable_events=True, visible=self.background is not None),
                 sg.Check('Hide legend', key='hide_legend', enable_events=True)]
            ]
        )

        r_sq, r_sq_adj = f_utils.r_squared_model_result(fit_result)
        results_text = (f'Fit Converged: {fit_result.success}\nR\u00B2: {r_sq:.5f}\n'
                        f'adjusted R\u00B2: {r_sq_adj:.5f}\n\n{fit_result.fit_report()}')

        results_tab = sg.Tab(
            'Results', [[sg.Multiline(results_text, disabled=True, key='results_output')]]
        )

        layout = [
            [sg.TabGroup([[plot_tab, results_tab]],
                         tab_background_color=sg.theme_background_color())],
            [sg.Column([
                [sg.Text('Accept fit results? '),
                 sg.Button('No'),
                 sg.Button('Yes', button_color=utils.PROCEED_COLOR)]
            ], element_justification='right', justification='right')]
        ]

        # alpha_channel=0 to make the window invisible until calling self.window.reappear()
        self.window = sg.Window('Fit Results', layout, finalize=True, alpha_channel=0, icon=utils._LOGO)
        self.window['results_output'].expand(expand_x=True, expand_y=True)


    def event_loop(self):
        """
        Handles the event loop for the GUI.

        Returns
        -------
        bool
            Returns True if the Continue button was pressed, otherwise False.

        """

        self.window.reappear()
        while True:
            event, values = self.window.read()
            if event in (sg.WIN_CLOSED, 'Yes', 'No'):
                break
            elif event == 'hide_legend':
                if values[event]:
                    self.axis.get_legend().set_visible(False)
                else:
                    self.axis.get_legend().set_visible(True)
                self.figure.canvas.draw_idle()
            elif event == 'subtract_bkg':
                if values[event]:
                    for line in self.axis.get_lines():
                        line.set_ydata(line.get_ydata() - self.background)
                    line.remove() # removes last line, which is the background
                else:
                    for line in self.axis.get_lines():
                        line.set_ydata(line.get_ydata() + self.background)
                    self.axis.plot(self.x, self.background, 'k-', label='background')

                self.axis.legend(ncol=max(1, len(self.axis.lines) // 4))
                if values['hide_legend']:
                    self.axis.get_legend().set_visible(False)
                self.axis.relim()
                self.axis.autoscale()
                self.figure.canvas.draw_idle()

        self._close()

        return event == 'Yes'


def _plot_model_component(axis, x, y, label):
    """
    Used to wrap the plotting of components in a try-except loop.

    Some models may not have the same size and the x-values, and
    thus would cause issues when plotting, so this function is used
    to avoid completely losing the results.

    Parameters
    ----------
    axis : plt.Axes
        The axis to plot the values on.
    x : array-like
        The x-values of the fit
    y : array-like
        The y-value(s) for the model.
    label : str
        The label to place on the plot for the component. Any '_' in
        the name is replace by ' ' and any leading or trailing whitespace
        is removed, to make the label appear nice.

    """

    try:
        axis.plot(x, y, label=label.replace('_', ' ').strip())
    except:
        print((f'Issue plotting the model result for component {label},'
               ' so it was not placed onto the figure.'))


def _find_peaks(dataframe, gui_values):
    """
    Finds peaks in the data according to the gui_values.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The dataframe that contains the x and y data.
    gui_values : dict
        A dictionary of values needed for finding the peaks.

    Returns
    -------
    found_peaks : list
        The list of peaks found in the data according to the
        peak finding parameters in gui_values.

    """

    x_data = utils.series_to_numpy(dataframe.iloc[:, gui_values['x_fit_index']])
    y_data = utils.series_to_numpy(dataframe.iloc[:, gui_values['y_fit_index']])
    nan_mask = (~np.isnan(x_data)) & (~np.isnan(y_data))
    x_min = max(gui_values['x_min'], np.nanmin(x_data))
    x_max = min(gui_values['x_max'], np.nanmax(x_data))

    additional_peaks = np.array(gui_values['peak_list'])
    additional_peaks = additional_peaks[(additional_peaks > x_min)
                                        & (additional_peaks < x_max)]

    found_peaks = peak_fitting.find_peak_centers(
        x_data[nan_mask], y_data[nan_mask], additional_peaks,
        gui_values['height'], gui_values['prominence'], x_min, x_max
    )[-1]

    return found_peaks


def _get_background_kwargs(gui_values):
    """
    Gets any necessary keyword arguments for the selected background model.

    Parameters
    ----------
    bkg_model : str
        The selected background model

    Returns
    -------
    kwargs : dict or None
        A dictionary containing any keyword arguments needed to initialize
        the selected background model. Returns None if no kwargs are required.

    """

    model = f_utils.get_model_name(gui_values['bkg_type'])

    if model not in f_utils._INIT_MODELS:
        kwargs = None
    else:
        total_kwargs = f_utils._INIT_MODELS[model]
        key, values = list(total_kwargs.items())[0] # only 1 item in the dict
        kwargs = {key: type(values[0])(gui_values[f'bkg_kwarg_{model}'])}

    return kwargs


def _create_peak_fitting_gui(default_inputs):
    """
    Creates the peak fitting portion of the fitting GUI.

    Parameters
    ----------
    user_inputs : dict
        A dictionary of values to construct the GUI.

    Returns
    -------
    layout : list(list(sg.Element))
        A list of lists of elements for the peak fitting portion
        of the fitting GUI.

    """

    disable_vary_voigt = True
    for model in itertools.chain(default_inputs['model_list'], default_inputs['default_model']):
        try:
            if f_utils.get_model_name(model) in ('VoigtModel', 'SkewedVoigtModel'):
                disable_vary_voigt = False
                break
        except KeyError:
            pass

    automatic_layout = [
        [sg.Text('Peak x values, separated by commas (leave blank to just use peak finding):')],
        [sg.Input(', '.join(str(val) for val in default_inputs['peak_list']),
                    key='peak_list', size=(50, 1))],
        [sg.Text('Prominence:', size=(13, 1)),
            sg.Input(default_inputs['prominence'], key='prominence', size=(5, 1))],
        [sg.Text('Minimum height:', size=(13, 1)),
            sg.Input(default_inputs['height'], key='height', size=(5, 1))],
        [sg.Text('Model list, separated by commas (leave blank to just use default model):')],
        [sg.Input(', '.join(str(val) for val in default_inputs['model_list']),
                    key='model_list', size=(50, 1), enable_events=True)]
    ]
    peak_finding_layout = sg.TabGroup(
        [
            [sg.Tab('Options', automatic_layout, key='automatic_tab',
                    visible=default_inputs['automatic_peaks']),
                sg.Tab('Options', [
                    [sg.Text('')],
                    [sg.Button('Launch Peak Selector', enable_events=True, size=(30, 5))]
                ], key='manual_tab', visible=default_inputs['manual_peaks'])]
        ], key='tab'
    )

    # models that require a keyword input during initialization
    bkg_init_models = []
    for model, kwargs in f_utils._INIT_MODELS.items():
        if model in f_utils._TOTAL_MODELS:
            key, values = list(kwargs.items())[0]
            bkg_init_models.append(
                sg.Column([
                    [sg.Text(f'Parameter: {key}'),
                     sg.Combo(values[1], default_inputs[f'bkg_kwarg_{model}'],
                              key=f'bkg_kwarg_{model}', readonly=True)]
                ], pad=(0, 0), key=f'bkg_col_{model}',
                visible=f_utils.get_model_name(default_inputs['bkg_type']) == model)
            )

    all_models = sorted(f_utils._GUI_MODELS.keys())
    peak_models = [f_utils.get_gui_name(model) for model in peak_fitting._PEAK_TRANSFORMS]
    auto_bkg_layout = [
        [sg.Text('Model'),
         sg.Combo(all_models, default_inputs['bkg_type'], key='bkg_type',
                  readonly=True, enable_events=True)],
        [sg.Text('    '), *bkg_init_models],
        [sg.Text('Min and max x values to use for fitting the background:')],
        [sg.Text('    x min:', size=(8, 1)),
            sg.Input(default_inputs['bkg_x_min'], key='bkg_x_min', size=(5, 1))],
        [sg.Text('    x max:', size=(8, 1)),
            sg.Input(default_inputs['bkg_x_max'], key='bkg_x_max', size=(5, 1))]
    ]

    bkg_layout = sg.TabGroup(
        [
            [sg.Tab('Background Options', auto_bkg_layout, key='automatic_bkg_tab',
                    visible=default_inputs['automatic_bkg']),
                sg.Tab('Background Options', [
                    [sg.Text('')],
                    [sg.Button('Launch Background Selector', key='bkg_selector',
                            enable_events=True, size=(30, 5))]
                ], key='manual_bkg_tab', visible=default_inputs['manual_bkg'])]
        ], key='bkg_tabs'
    )

    layout = [
        [sg.Text('Default peak model:'),
         sg.Combo(peak_models, key='default_model', readonly=True,
                    default_value=default_inputs['default_model'], enable_events=True)],
        [sg.Check('Vary Voigt gamma parameter', key='vary_voigt', disabled=disable_vary_voigt,
                    default=default_inputs['vary_voigt'],
                    tooltip='if True, will allow the gamma parameter in the Voigt model'\
                            ' to be varied as an additional variable')],
        [sg.Text('Peak full-width at half-maximum:'),
            sg.Input(default_inputs['peak_width'], key='peak_width', size=(5, 1))],
        [sg.Text('Center offset:'),
            sg.Input(default_inputs['center_offset'], key='center_offset', size=(5, 1))],
        [sg.Text('Maximum sigma value:'),
            sg.Input(default_inputs['max_sigma'], key='max_sigma', size=(5, 1))],
        [sg.Check('Fit residuals', key='fit_residuals', enable_events=True,
                    default=default_inputs['fit_residuals'])],
        [sg.Text('Minimum residual height:'),
            sg.Input(default_inputs['min_resid'], key='min_resid', size=(5, 1),
                    visible=default_inputs['fit_residuals'])],
        [sg.Text('Number of residual fits:'),
            sg.Input(default_inputs['num_resid_fits'], key='num_resid_fits',
                    size=(5, 1), visible=default_inputs['fit_residuals'])],
        [sg.Text('')],
        [sg.Check('Subtract background', key='subtract_bkg', enable_events=True,
                    default=default_inputs['subtract_bkg']),
            sg.Text('('),
            sg.Radio('Automatic', 'bkg_fitting', key='automatic_bkg',
                    enable_events=True, default=default_inputs['automatic_bkg']),
            sg.Radio('Manual', 'bkg_fitting', key='manual_bkg',
                    enable_events=True, default=default_inputs['manual_bkg']),
            sg.Text(')')],
        [bkg_layout],
        [sg.Text('Peak Finding Options', relief='ridge', size=(60, 1),
                    pad=(5, (20, 10)), justification='center')],
        [sg.Radio('Automatic Peak Finding', 'peak_finding', key='automatic_peaks',
                    enable_events=True, default=default_inputs['automatic_peaks']),
            sg.Radio('Manual Peak Finding', 'peak_finding', key='manual_peaks',
                    enable_events=True, default=default_inputs['manual_peaks'],
                    pad=((50, 10), 5))],
        [peak_finding_layout],
    ]

    return layout


def _create_general_fitting_gui(user_inputs):
    """
    Creates the general fitting portion of the fitting GUI.

    Parameters
    ----------
    user_inputs : dict
        A dictionary of values to construct the GUI.

    Returns
    -------
    layout : list(list(sg.Element))
        A list of lists of elements for the general fitting portion
        of the fitting GUI.

    """


def _create_fitting_gui(dataframe, user_inputs=None):
    """
    Creates the fitting gui.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The dataframe with the data to fit.
    user_inputs : dict, optional
        A dictionary of values to override the default inputs for the GUI.

    Returns
    -------
    window : sg.Window
        The fitting GUI.
    default_inputs : dict
        A dictionary of the values used to create the window.

    """

    # The default inputs for the GUI
    default_inputs = {
        'sample_name': 'Sample',
        'x_fit_index': '0',
        'y_fit_index': '1',
        'x_label': 'raw x data',
        'y_label': 'raw y data',
        'x_min': '-inf',
        'x_max': 'inf',
        'min_method': 'least_squares',
        'batch_fit': False,
        'peak_list': [],
        'prominence': 'inf',
        'height': '-inf',
        'model_list': [],
        'default_model': 'Gaussian',
        'vary_voigt': False,
        'peak_width': '',
        'center_offset': '',
        'max_sigma': 'inf',
        'subtract_bkg': True,
        'bkg_type': 'Constant',
        'bkg_x_min': '-inf',
        'bkg_x_max': 'inf',
        'fit_residuals': False,
        'min_resid': '0.05',
        'num_resid_fits': '5',
        'automatic_peaks' : True,
        'manual_peaks': False,
        'debug': False,
        'automatic_bkg': True,
        'manual_bkg': False,
        'confirm_results': True,
        'selected_peaks': [],
        'selected_bkg': []
    }

    for model, kwargs in f_utils._INIT_MODELS.items():
        default_inputs[f'bkg_kwarg_{model}'] = list(kwargs.values())[0][0]

    if user_inputs is not None:
        default_inputs.update(user_inputs)

    tab1 = _create_peak_fitting_gui(default_inputs)
    #tab2 = _create_general_fitting_gui(user_inputs) #TODO this is for later

    column_layout = [
        [sg.Text('Raw Data', relief='ridge', size=(60, 1),
                    justification='center')],
        [sg.Text('Sample Name:'),
            sg.Input(default_inputs['sample_name'], key='sample_name', size=(20, 1))],
        [sg.Text('Column of x data for fitting:'),
            sg.Combo(list(range(len(dataframe.columns))), size=(3, 1), readonly=True,
                    key='x_fit_index', default_value=default_inputs['x_fit_index'])],
        [sg.Text('Column of y data for fitting:'),
            sg.Combo(list(range(len(dataframe.columns))), size=(3, 1), readonly=True,
                    key='y_fit_index', default_value=default_inputs['y_fit_index'])],
        [sg.Text('x data label:'),
            sg.Input(default_inputs['x_label'], key='x_label', size=(20, 1))],
        [sg.Text('y data label:'),
            sg.Input(default_inputs['y_label'], key='y_label', size=(20, 1))],
        [sg.Text('Min and max values to use for fitting:')],
        [sg.Text('    x min:', size=(8, 1)),
            sg.Input(default_inputs['x_min'], key='x_min', size=(5, 1))],
        [sg.Text('    x max:', size=(8, 1)),
            sg.Input(default_inputs['x_max'], key='x_max', size=(5, 1))],
        [sg.Text('Minimization method:'),
            sg.Combo(['least_squares', 'leastsq'], key='min_method', readonly=False,
                    default_value=default_inputs['min_method'])],
        [sg.Text('Fitting Options', relief='ridge', size=(60, 1),
                    pad=(5, (20, 10)), justification='center')],
        *tab1
    ]

    layout = [
        [sg.Frame('', [
            [sg.Column(column_layout, scrollable=True,
                       vertical_scroll_only=True, size=(700, 500))]
            ])],
        [sg.Check('Batch Fit', default_inputs['batch_fit'], key='batch_fit')],
        [sg.Check('Confirm Fit Results', default_inputs['confirm_results'],
                    key='confirm_results')],
        [sg.Check('Debug Fitting Process', default_inputs['debug'], key='debug')],
        [sg.Text('')],
        [sg.Button('Fit', bind_return_key=True, size=(6, 1),
                    button_color=utils.PROCEED_COLOR),
         sg.Button('Test Plot'),
         sg.Button('Show Data'),
         sg.Button('Reset to Default'),
         sg.Button('Skip Fitting')]
    ]

    window = sg.Window('Fitting', layout, finalize=True, icon=utils._LOGO)
    if default_inputs['manual_peaks']:
        window['manual_tab'].select()
        window['automatic_tab'].update(visible=False)
    if default_inputs['manual_bkg']:
        window['manual_bkg_tab'].select()
        window['automatic_bkg_tab'].update(visible=False)

    return window, default_inputs


def _process_fit_results(fit_result):
    """
    Gets the individual models and parameters from a fit result.

    Parameters
    ----------
    fit_result : lmfit.ModelResult
        The ModelResult object from the fitting.

    Returns
    -------
    params_dataframe : pd.DataFrame
        The dataframe containing the parameter values and standard errors
        for each model.
    models_data : list(pd.DataFrame)
        The list of dataframes containing the values for each individual
        model within the fit result.

    """

    # Creation of dataframe for best values of all model parameters
    individual_models = {}
    best_values = {}
    for (prefix, model_values), model in zip(fit_result.eval_components().items(), fit_result.components):
        model_name = model.__class__.__name__
        # use the GUI name for built-in models
        try:
            best_values[prefix] = {'Model': f_utils.get_gui_name(model_name)}
        except KeyError:
            best_values[prefix] = {'Model': model_name}
        individual_models[prefix] = f_utils._check_if_constant(
            model_name, model_values, fit_result.data
        )
        for param in model.param_names:
            param_ = fit_result.params[param]
            best_values[prefix][f'{param[len(prefix):]}__VALUE__'] = param_.value
            best_values[prefix][f'{param[len(prefix):]}__STDERR__'] = (
                param_.stderr if param_.stderr not in (None, np.nan) else 'N/A')

        if f_utils.get_is_peak(model_name):
            numeric_calcs = {
                'numeric area': f_utils.numerical_area(fit_result.userkws['x'], model_values),
                'numeric fwhm': f_utils.numerical_fwhm(fit_result.userkws['x'], model_values),
                'numeric extremum': f_utils.numerical_extremum(model_values),
                'numeric mode': f_utils.numerical_mode(fit_result.userkws['x'], model_values)
            }
            for calc, calc_value in numeric_calcs.items():
                best_values[prefix][f'{calc}__VALUE__'] = calc_value if calc_value is not None else 'N/A'
                best_values[prefix][f'{calc}__STDERR__'] = 'None'

    params_dataframe = pd.DataFrame(best_values).transpose().fillna('-')
    params_dataframe.index = [index.replace('_', ' ').strip() for index in params_dataframe.index]

    # Creation of dataframe for model values
    models_data = []
    bkg_term = ' + background' if 'background_' in individual_models else ''
    bkg = individual_models.get('background_', 0)
    for term, value in individual_models.items():
        key = term.replace('_', ' ').strip()
        if term != 'background_':
            key += bkg_term
            data = value + bkg
        else:
            data = value
        try:
            models_data.append(pd.DataFrame({key: data}))
        except ValueError: # data is scalar or np.array with size == 1
            models_data.append(pd.DataFrame({key: [data]}))
    models_data.append(pd.DataFrame({'total fit': fit_result.best_fit}))

    return params_dataframe, models_data


def _process_fitting_kwargs(dataframe, values):
    """
    Processes the values from the fitting GUI to fit data and output the results.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The dataframe with the data to fit.
    values : dict
        A dictionary of values determining the processing steps from the fitting GUI.

    Returns
    -------
    fit_result : lmfit.model.ModelResult or None
        A lmfit.ModelResult object, which gives information for the fit
        done on the dataset. Is None if fitting was skipped.
    values_df : pd.DataFrame or None
        The dataframe containing the x and y data, the y data
        for every individual model, the summed y data of all models,
        and the background, if present. Is None if fitting was skipped.
    params_df : pd.DataFrame or None
        The dataframe containing the value and standard error
        associated with all of the parameters in the fitting
        (eg. coefficients for the baseline, areas and sigmas for each peak).
        Is None if fitting was skipped.
    descriptors_df : pd.DataFrame or None
        The dataframe which contains some additional information about the
        fitting. Currently has the adjusted r squared, reduced chi squared,
        the Akaike information criteria, the Bayesian information criteria,
        and the minimization method used for fitting. Is None if fitting was
        skipped.

    """

    x_label = values['x_label']
    y_label = values['y_label'] if values['y_label'] != x_label else values['y_label'] + '_1'
    x_data = utils.series_to_numpy(dataframe.iloc[:, values['x_fit_index']])
    y_data = utils.series_to_numpy(dataframe.iloc[:, values['y_fit_index']])
    x_min = values['x_min']
    x_max = values['x_max']
    default_model = values['default_model']
    vary_voigt = values['vary_voigt']
    center_offset = values['center_offset']
    min_method = values['min_method']
    subtract_bkg = values['subtract_bkg']
    background_type = values['bkg_type']
    background_kwargs = _get_background_kwargs(values)
    bkg_min = values['bkg_x_min']
    bkg_max = values['bkg_x_max']
    max_sigma = values['max_sigma']
    fit_residuals = values['fit_residuals']
    min_resid = values['min_resid']
    num_resid_fits = values['num_resid_fits']
    debug = values['debug']

    if values['manual_peaks']:
        peaks = sorted(values['selected_peaks'], key=lambda x: x[3])
        model_list = [peak[0] for peak in peaks]
        peak_heights = [peak[1] for peak in peaks]
        peak_width = [peak[2] for peak in peaks]
        additional_peaks = [peak[3] for peak in peaks]
        # ensures no additional peaks are found
        height = np.inf
        prominence = np.inf

    else:
        additional_peaks = values['peak_list']
        model_list = values['model_list']
        peak_width = values['peak_width']
        peak_heights = None
        height = values['height']
        prominence = values['prominence']

    if subtract_bkg and values['manual_bkg']:
        subtract_bkg = False
        y_data = f_utils.subtract_linear_background(x_data, y_data, values['selected_bkg'])

    #TODO make all but x_data and y_data into keyword arguments
    fitting_results = peak_fitting.fit_peaks(
        x_data, y_data, height, prominence, center_offset, peak_width, default_model,
        subtract_bkg, bkg_min, bkg_max, 0, max_sigma, min_method, x_min, x_max,
        additional_peaks, model_list, background_type, background_kwargs, None,
        vary_voigt, fit_residuals, num_resid_fits, min_resid, debug, peak_heights
    )

    # only take the last fit from the list
    fit_result = fitting_results['fit_results'][-1]
    params_df, models_data = _process_fit_results(fit_result)

    # Creation of dataframe for raw data and peak values
    # Raw data
    total_data = [pd.DataFrame({
        x_label: dataframe.iloc[:, values['x_fit_index']],
        y_label: dataframe.iloc[:, values['y_fit_index']]
    })]
    # Data used for fitting
    total_data.extend([
        pd.DataFrame({x_label: fit_result.userkws['x']}),
        pd.DataFrame({y_label: fit_result.data})
    ])
    total_data.extend(models_data)

    # use concat with DataFrames to ensure that the sizing is correct
    # even if the sizes of individual components do not match; also
    # allows columns to have repeated names
    values_df = pd.concat(total_data, axis=1)

    # Creation of dataframe for descriptions of the fitting
    r_sq, adj_r_sq = f_utils.r_squared_model_result(fit_result)
    # use fit_result.data.size for data points b/c when a fit fails, fit_result.ndata is
    # set to 1; same reason the degrees of freedom is not set to fit_result.nfree
    descriptors_df = pd.DataFrame(
        [fit_result.success, r_sq, adj_r_sq, fit_result.chisqr,
         fit_result.redchi, fit_result.aic, fit_result.bic, min_method,
         fit_result.data.size, fit_result.nvarys],
        index=['Fit converged', 'R\u00B2', 'Adjusted R\u00B2', '\u03c7\u00B2',
               'Reduced \u03c7\u00B2', 'A.I.C.', 'B.I.C.', 'Minimization method',
               'Data points', 'Independant variables']
    )

    return fit_result, values_df, params_df, descriptors_df


def fit_dataframe(dataframe, user_inputs=None):
    """
    Creates a GUI to select data from a dataframe for fitting.

    Parameters
    ----------
    dataframe : pd.DataFrame
        A pandas dataframe containing the data to be fit.
    user_inputs : dict, optional
        Values to use as the default inputs in the GUI.

    Returns
    -------
    fit_result : lmfit.model.ModelResult or None
        A lmfit.ModelResult object, which gives information for the fit
        done on the dataset. Is None if fitting was skipped.
    values_df : pd.DataFrame or None
        The dataframe containing the x and y data, the y data
        for every individual model, the summed y data of all models,
        and the background, if present. Is None if fitting was skipped.
    params_df : pd.DataFrame or None
        The dataframe containing the value and standard error
        associated with all of the parameters in the fitting
        (eg. coefficients for the baseline, areas and sigmas for each peak).
        Is None if fitting was skipped.
    descriptors_df : pd.DataFrame or None
        The dataframe which contains some additional information about the
        fitting. Currently has the adjusted r squared, reduced chi squared,
        the Akaike information criteria, the Bayesian information criteria,
        and the minimization method used for fitting. Is None if fitting was
        skipped.
    gui_values : dict or None
        The values selected in the GUI for all of the various fields, which
        can be used to reuse the values from a past interation. Is None if
        fitting was skipped.

    """

    if user_inputs is not None and user_inputs.get('batch_fit', False):
        gui_values = user_inputs.copy()
    else:
        gui_values = _fitting_gui_event_loop(dataframe, user_inputs)

    fit_result = values_df = params_df = descriptors_df = None
    while gui_values is not None:
        try:
            fit_result, values_df, params_df, descriptors_df = _process_fitting_kwargs(dataframe, gui_values)
        except Exception:
            sg.popup(f'Error occurred during fitting:\n{traceback.format_exc()}\n', icon=utils._LOGO)
            gui_values = _fitting_gui_event_loop(dataframe, gui_values)
        else:
            if gui_values['confirm_results'] and not ResultsPlot(fit_result).event_loop():
                gui_values = _fitting_gui_event_loop(dataframe, gui_values)
            else:
                break

    return fit_result, values_df, params_df, descriptors_df, gui_values


def _fitting_gui_event_loop(dataframe, user_inputs):
    """
    Handles the event loop for the fitting gui.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The dataframe with the data to fit.
    user_inputs : dict, optional
        A dictionary of inputs , by default None

    Returns
    -------
    values : dict or None
        A dictionary of values from the GUI after the event loop. If
        fitting for the dataframe was skipped, values is None.

    """

    validations = {
        'peak_fitting': {
            'integers': [
                ['x_fit_index', 'x column'],
                ['y_fit_index', 'y column'],
                ['num_resid_fits', 'number of residual fits']
            ],
            'floats': [
                ['x_min', 'x min'],
                ['x_max', 'x max'],
                ['bkg_x_min', 'background x min'],
                ['bkg_x_max', 'background x max'],
                ['peak_width', 'peak width'],
                ['height', 'minimum height'],
                ['prominence', 'prominence'],
                ['center_offset', 'center offset'],
                ['min_resid', 'minimum residual height'],
                ['max_sigma', 'maximum sigma']
            ],
            'strings': [
                ['bkg_type', 'background model'],
                ['min_method', 'minimization method'],
                ['default_model', 'default model'],
            ],
            'user_inputs': [
                ['peak_list', 'peak x values', float, True],
                ['sample_name', 'sample name', utils.string_to_unicode, False, None],
                ['sample_name', 'sample name', utils.validate_sheet_name, False, None],
                ['x_label', 'x label', utils.string_to_unicode, False, None],
                ['y_label', 'y label', utils.string_to_unicode, False, None],
                ['model_list', 'model list', str, True],
            ],
            'constraints': [
                ['peak_width', 'peak width', '> 0'],
                ['center_offset', 'center offset', '>= 0'],
                ['max_sigma', 'maximum sigma', '> 0'],
                ['num_resid_fits', 'number of residual fits', '> 0']
            ]
        }
    }

    validations['plotting'] = {
        'integers': validations['peak_fitting']['integers'][:2],
        'floats': validations['peak_fitting']['floats'][:4]
                  + validations['peak_fitting']['floats'][5:7],
        'user_inputs': validations['peak_fitting']['user_inputs'][:1],
    }
    validations['peak_selector'] = {
        'integers': validations['peak_fitting']['integers'][:2],
        'floats': validations['peak_fitting']['floats'][:5],
        'strings': validations['peak_fitting']['strings'][:1],
        'constraints': validations['peak_fitting']['constraints'][:1]
    }
    validations['bkg_selector'] = {
        'integers': validations['peak_fitting']['integers'][:2],
    }

    peak_models = peak_fitting._PEAK_TRANSFORMS
    voigt_models = [f_utils.get_gui_name(model) for model in ('VoigtModel', 'SkewedVoigtModel')]

    window, default_inputs = _create_fitting_gui(dataframe, user_inputs)
    peak_list = default_inputs['selected_peaks'] # Values if using manual peak selection
    bkg_points = default_inputs['selected_bkg'] # Values if using manual background selection
    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED:
            utils.safely_close_window(window)

        elif event == 'Skip Fitting':
            skip = sg.popup_yes_no(
                'Peak fitting will be skipped for this entry.\n\nProceed?\n',
                title='Skip Fitting', icon=utils._LOGO
            )
            if skip == 'Yes':
                values = None
                break

        elif event == 'Reset to Default':
            reset = sg.popup_yes_no(
                        'All values will be returned to their default.\n\nProceed?\n',
                        title='Reset to Defaults', icon=utils._LOGO
                    )
            if reset == 'Yes':
                defaults = {
                    key: vals for key, vals in default_inputs.items()
                    if key not in {'selected_peaks', 'selected_bkg', 'bkg_tabs', 'tab'}
                }
                window.fill(defaults)
                peak_list = []
                bkg_points = []

        elif (event == 'Test Plot'
                and utils.validate_inputs(values, **validations['plotting'])):
            window.hide()
            try:
                SimpleEmbeddedFigure(dataframe, values).event_loop()
            except Exception as e:
                sg.popup(f'Error creating plot:\n    {repr(e)}', icon=utils._LOGO)

            window.un_hide()

        elif event == 'Show Data':
            data_window = utils.show_dataframes(dataframe)
            if data_window is not None: #TODO check if this is still needed
                data_window.finalize().TKroot.grab_set()
                data_window.read(close=True)
                data_window = None

        elif (event == 'bkg_selector'
                and utils.validate_inputs(values, **validations['bkg_selector'])):
            window.hide()

            x_data = utils.series_to_numpy(dataframe.iloc[:, values['x_fit_index']])
            y_data = utils.series_to_numpy(dataframe.iloc[:, values['y_fit_index']])
            try:
                bkg_points = peak_fitting.BackgroundSelector(
                    x_data, y_data, bkg_points).event_loop()
            except Exception as e:
                sg.popup(f'Error launching Background Selector:\n    {repr(e)}', icon=utils._LOGO)

            window.un_hide()

        elif (event == 'Launch Peak Selector'
                and utils.validate_inputs(values, **validations['peak_selector'])):
            window.hide()

            x_data = utils.series_to_numpy(dataframe.iloc[:, values['x_fit_index']])
            y_data = utils.series_to_numpy(dataframe.iloc[:, values['y_fit_index']])
            x_min = values['x_min']
            x_max = values['x_max']
            bkg_min = values['bkg_x_min']
            bkg_max = values['bkg_x_max']
            subtract_bkg = values['subtract_bkg']
            background_type = values['bkg_type']
            background_kwargs = _get_background_kwargs(values)
            bkg_min = values['bkg_x_min']
            bkg_max = values['bkg_x_max']
            peak_width = values['peak_width']
            default_model = values['default_model']

            if subtract_bkg and values['manual_bkg']:
                subtract_bkg = False
                y_data = f_utils.subtract_linear_background(x_data, y_data, bkg_points)

            domain_mask = (x_data >= x_min) & (x_data <= x_max)
            try:
                peak_list = peak_fitting.PeakSelector(
                    x_data[domain_mask], y_data[domain_mask], peak_list,
                    peak_width, subtract_bkg, background_type,
                    background_kwargs, bkg_min, bkg_max, default_model
                ).event_loop()
            except Exception as e:
                sg.popup(f'Error launching Peak Selector:\n    {repr(e)}', icon=utils._LOGO)
            else:
                # updates values in the window from the peak selector plot
                sorted_peaks = [[val[0], val[3]] for val in sorted(peak_list, key=lambda x: x[3])]
                temp_model_list = [f_utils.get_gui_name(model) for model, _ in sorted_peaks]
                window['model_list'].update(value=', '.join(temp_model_list))
                window['peak_list'].update(
                    value=', '.join([str(np.round(center, 2)) for _, center in sorted_peaks])
                )

                if any(model in voigt_models for model in temp_model_list):
                    window['vary_voigt'].update(disabled=False)
                elif values['default_model'] not in voigt_models:
                    window['vary_voigt'].update(disabled=True, value=False)

            window.un_hide()
        # toggle auto/manual peak selection
        elif event in ('automatic_peaks', 'manual_peaks'):
            if event == 'automatic_peaks':
                window['automatic_tab'].update(visible=True)
                window['automatic_tab'].select()
                window['manual_tab'].update(visible=False)
            else:
                window['automatic_tab'].update(visible=False)
                window['manual_tab'].update(visible=True)
                window['manual_tab'].select()
        # toggle auto/manual background selection
        elif event in ('automatic_bkg', 'manual_bkg'):
            if event == 'automatic_bkg':
                window['automatic_bkg_tab'].update(visible=True)
                window['automatic_bkg_tab'].select()
                window['manual_bkg_tab'].update(visible=False)
            else:
                window['manual_bkg_tab'].update(visible=True)
                window['manual_bkg_tab'].select()
                window['automatic_bkg_tab'].update(visible=False)

        elif event == 'subtract_bkg':
            if values['subtract_bkg']:
                window['bkg_type'].update(visible=True)
                window['bkg_x_min'].update(visible=True)
                window['bkg_x_max'].update(visible=True)
                window['automatic_bkg'].update(disabled=False)
                window['manual_bkg'].update(disabled=False)
                window['bkg_selector'].update(disabled=False)
            else:
                # set bkg to 'Constant' b/c it has no init kwargs and its GUI name will not change
                window['bkg_type'].update(visible=False, value='Constant')
                window['bkg_x_min'].update(visible=False, value='-inf')
                window['bkg_x_max'].update(visible=False, value='inf')
                window['automatic_bkg'].update(disabled=True)
                window['manual_bkg'].update(disabled=True)
                window['bkg_selector'].update(disabled=True)
                for model in f_utils._INIT_MODELS.keys():
                    if model in f_utils._TOTAL_MODELS:
                        window[f'bkg_col_{model}'].update(visible=False)

        elif event == 'bkg_type':
            bkg_model = f_utils.get_model_name(values['bkg_type'])
            for model in f_utils._INIT_MODELS.keys():
                if model == bkg_model:
                    window[f'bkg_col_{model}'].update(visible=True)
                elif model in f_utils._TOTAL_MODELS:
                    window[f'bkg_col_{model}'].update(visible=False)

        elif event in ('model_list', 'default_model'):
            temp_model_list = []
            for entry in values['model_list'].split(','):
                if entry:
                    try:
                        temp_model_list.append(f_utils.get_gui_name(entry.strip()))
                    except KeyError:
                        pass
            if (any(model in voigt_models for model in temp_model_list)
                    or values['default_model'] in voigt_models):
                window['vary_voigt'].update(disabled=False)
            else:
                window['vary_voigt'].update(disabled=True, value=False)

        elif event == 'fit_residuals':
            if values['fit_residuals']:
                window['min_resid'].update(visible=True)
                window['num_resid_fits'].update(visible=True)
            else:
                window['min_resid'].update(visible=False, value=0.05)
                window['num_resid_fits'].update(visible=False, value=5)

        elif event == 'Fit' and utils.validate_inputs(values, **validations['peak_fitting']):
            bad_models = []
            for entry in values['model_list']:
                try:
                    if f_utils.get_model_name(entry) not in peak_models:
                        bad_models.append(entry)
                except KeyError:
                    bad_models.append(entry)
            if bad_models:
                sg.popup(
                    f'Need to correct terms in the model list:\n  {", ".join(bad_models)}\n',
                    title='Error', icon=utils._LOGO
                )
            elif values['automatic_peaks'] and not _find_peaks(dataframe, values):
                sg.popup(
                    ('No peaks found in fitting range. Either manually enter \n'
                        'peak positions or change peak finding options.\n'),
                    title='Error', icon=utils._LOGO
                )
            elif values['manual_peaks'] and not peak_list:
                sg.popup(
                    'Please manually select peaks or change peak finding to automatic.\n',
                    title='Error', icon=utils._LOGO
                )
            else:
                break

    window.close()
    del window

    if values is not None:
        values['selected_peaks'] = peak_list
        values['selected_bkg'] = bkg_points

    return values


def fit_to_excel(values_dataframe, params_dataframe, descriptors_dataframe,
                 excel_writer_handler, sheet_name=None, plot_excel=False):
    """
    Outputs the relevant data from peak fitting to an Excel file.

    Parameters
    ----------
    values_dataframe : pd.DataFrame
        The dataframe containing the x and y data, the y data
        for every individual model, the summed y data of all models,
        and the background, if present.
    params_dataframe : pd.DataFrame
        The dataframe containing the value and standard error
        associated with all of the parameters in the fitting
        (eg. coefficients for the baseline, areas and sigmas for each peak).
    descriptors_dataframe : pd.DataFrame
        The dataframe which contains some additional information about the fitting.
        Currently has the adjusted r squared, reduced chi squared, the Akaike
        information criteria, the Bayesian information criteria, and the minimization
        method used for fitting.
    excel_writer_handler : mcetl.excel_writer.ExcelWriterHandler
        The ExcelWriterHandler that contains the pandas ExcelWriter object for
        writing to Excel, and the styles for styling the cells in Excel.
    sheet_name: str, optional
        The Excel sheet name.
    plot_excel : bool, optional
        If True, will create a simple plot in Excel that plots the raw x and
        y data, the data for each peak, the total of all peaks, and
        the background if it is present.

    """

    from openpyxl.chart import Reference, Series, ScatterChart
    from openpyxl.utils.dataframe import dataframe_to_rows

    excel_writer = excel_writer_handler.writer
    style_cache = excel_writer_handler.style_cache

    # Ensures that the sheet name is unique so it does not overwrite data;
    # not needed for openpyxl, but just a precaution
    current_sheets = [sheet.title.lower() for sheet in excel_writer.book.worksheets]
    if sheet_name is not None:
        sheet_name = utils.string_to_unicode(sheet_name)
        sheet_base = sheet_name
    else:
        sheet_base = 'Sheet'
        sheet_name = 'Sheet_1'
    num = 1
    while sheet_name.lower() in current_sheets:
        sheet_name = f'{sheet_base}_{num}'
        num += 1

    lengths = {
        'params': len(params_dataframe.columns),
        'values': len(values_dataframe.columns),
        'descriptors': len(descriptors_dataframe.columns)
    }
    lengths['total'] = sum(lengths.values()) + 1

    # use dict.fromkeys rather than a set to preserve order
    param_names = dict.fromkeys([
        '',
        *[name.replace('__STDERR__', '').replace('__VALUE__', '') for name in params_dataframe.columns]
    ])

    # Easier to just write params and descriptors using pandas rather than using
    # openpyxl; will not cost significant time since there are only a few cells
    params_dataframe.to_excel(
        excel_writer, sheet_name=sheet_name, startrow=3,
        startcol=lengths['values'], header=False, index=True
    )
    descriptors_dataframe.to_excel(
        excel_writer, sheet_name=sheet_name, startrow=1,
        startcol=lengths['values'] + lengths['params'] + 1,
        header=False, index=True
    )
    worksheet = excel_writer.book[sheet_name]

    # Write and format all headers
    headers = (
        {'name': 'Values', 'start': 1, 'end': lengths['values']},
        {'name': 'Model Parameters', 'start': lengths['values'] + 1,
         'end': lengths['values'] + lengths['params'] + 1},
        {'name': 'Fit Description',
         'start': lengths['values'] + lengths['params'] + 2,
         'end': lengths['total'] + 1}
    )
    suffix = itertools.cycle(['even', 'odd'])
    for header in headers:
        style_suffix = next(suffix)
        worksheet.merge_cells(
            start_row=1, start_column=header['start'], end_row=1, end_column=header['end']
        )
        worksheet.cell(row=1, column=header['start'], value=header['name'])
        for col in range(header['start'], header['end'] + 1):
            setattr(
                worksheet.cell(row=1, column=col),
                *style_cache['fitting_header_' + style_suffix]
            )

    # Subheaders for values_dataframe
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    setattr(worksheet.cell(row=2, column=1, value='Raw Data'), *style_cache['fitting_header_odd'])
    setattr(worksheet.cell(row=2, column=2), *style_cache['fitting_header_odd'])
    worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
    setattr(worksheet.cell(row=2, column=3, value='Fit Data'), *style_cache['fitting_header_even'])
    setattr(worksheet.cell(row=2, column=4), *style_cache['fitting_header_even'])
    worksheet.merge_cells(start_row=2, start_column=5, end_row=2,
                          end_column=lengths['values'])
    worksheet.cell(row=2, column=5, value='Fit Output')
    for col in range(5, lengths['values'] + 1):
        setattr(worksheet.cell(row=2, column=col), *style_cache['fitting_header_odd'])

    # Formatting for values_dataframe
    suffix = itertools.cycle(['even', 'odd'])
    for i, peak_name in enumerate(values_dataframe.columns, 1):
        setattr(
            worksheet.cell(row=3, column=i, value=peak_name),
            *style_cache['fitting_subheader_' + next(suffix)]
        )

    rows = dataframe_to_rows(values_dataframe, index=False, header=False)
    for row_index, row in enumerate(rows, 4):
        suffix = itertools.cycle(['even', 'odd'])
        for column_index, value in enumerate(row, 1):
            setattr(
                worksheet.cell(row=row_index, column=column_index, value=value),
                *style_cache['fitting_columns_' + next(suffix)]
            )

    # Formatting for params_dataframe
    for index, subheader in enumerate(param_names):
        style_suffix = next(suffix)

        if index < 2:
            prefix = 'fitting_columns_' if index == 0 else 'fitting_subheader_'
            column = lengths['values'] + 1 + index
            worksheet.merge_cells(
                start_row=2, start_column=column, end_row=3, end_column=column
            )
            setattr(
                worksheet.cell(row=2, column=column, value=subheader),
                *style_cache[prefix + style_suffix]
            )
            setattr(
                worksheet.cell(row=3, column=column),
                *style_cache[prefix + style_suffix]
            )
            prefix = 'fitting_descriptors_' if index == 0 else 'fitting_columns_'
            for row in range(len(params_dataframe)):
                setattr(
                    worksheet.cell(row=4 + row, column=column),
                    *style_cache[prefix + style_suffix]
                )
        else:
            column = lengths['values'] + 1 + (2 * (index - 1))
            worksheet.merge_cells(
                start_row=2, start_column=column, end_row=2, end_column=column + 1
            )
            setattr(
                worksheet.cell(row=2, column=column, value=subheader),
                *style_cache['fitting_subheader_' + style_suffix]
            )
            setattr(
                worksheet.cell(row=2, column=column + 1),
                *style_cache['fitting_subheader_' + style_suffix]
            )
            setattr(
                worksheet.cell(row=3, column=column, value='Value'),
                *style_cache['fitting_subheader_' + style_suffix]
            )
            setattr(
                worksheet.cell(row=3, column=column + 1, value='Standard Error'),
                *style_cache['fitting_subheader_' + style_suffix]
            )

            for row in range(len(params_dataframe)):
                setattr(
                    worksheet.cell(row=4 + row, column=column),
                    *style_cache['fitting_columns_' + style_suffix]
                )
                setattr(
                    worksheet.cell(row=4 + row, column=column + 1),
                    *style_cache['fitting_columns_' + style_suffix]
                )

    # Formatting for descriptors_dataframe
    for column in range(2):
        style = 'fitting_descriptors_' + next(suffix)
        for row in range(len(descriptors_dataframe.index)):
            setattr(
                worksheet.cell(
                    row=2 + row,
                    column=column + lengths['values'] + lengths['params'] + 2
                ),
                *style_cache[style]
            )

    # Adjust column and row dimensions
    worksheet.row_dimensions[1].height = 18
    for column in range(1, lengths['values'] + lengths['params'] + 2):
        worksheet.column_dimensions[utils.excel_column_name(column)].width = 12.5
    for column in range(lengths['values'] + lengths['params'] + 2, lengths['total'] + 2):
        worksheet.column_dimensions[utils.excel_column_name(column)].width = 20

    if plot_excel:
        axis_attributes = {
            'x_axis': {
                'title': values_dataframe.columns[0],
                'crosses': 'min'
            },
            'y_axis': {
                'title': values_dataframe.columns[1],
                'crosses': 'min'
            }
        }
        chart = ScatterChart()
        for axis, attribute in axis_attributes.items():
            for axis_attribute, value in attribute.items():
                setattr(getattr(chart, axis), axis_attribute, value)

        # plot everything but the raw data
        for i, peak_name in enumerate(values_dataframe.columns[3:], 4):
            chart.append(
                Series(
                    Reference(worksheet, i, 3, i, len(values_dataframe.index) + 3),
                    xvalues=Reference(worksheet, 3, 4, 3, len(values_dataframe.index) + 3),
                    title_from_data=True
                )
            )

        worksheet.add_chart(chart, 'D8')


def launch_fitting_gui(dataframe=None, gui_values=None, excel_writer=None,
                       save_excel=True, plot_excel=True, mpl_changes=None,
                       save_when_done=True, excel_formats=None):
    """
    Convenience function to fit dataframe(s) and write their results to Excel.

    Parameters
    ----------
    dataframe : pd.DataFrame or list/tuple, optional
        The dataframe or list/tuple of dataframes to fit.
    gui_values : dict, optional
        A dictionary containing the default gui values to pass to fit_dataframe.
    excel_writer : pd.ExcelWriter, optional
        The Excel writer used to save the results to Excel. If input, the engine
        must be "openpyxl".
    save_excel : bool, optional
        If True (default), then the fit results will be saved to an Excel file.
    plot_excel : bool, optional
        If True (default), then the fit results will be plotted in the
        Excel file (if saving).
    mpl_changes : dict, optional
        A dictionary of changes to apply to matplotlib's rcParams file, which
        affects how plots look.
    save_when_done : bool, optional
        If True (default), then the Excel file will be saved once all dataframes
        are fit.
    excel_formats : dict, optional
        A dictionary of formats to use when writing to Excel. The dictionary must
        have one of the following keys:
            'fitting_header_even', 'fitting_header_odd', 'fitting_subheader_even',
            'fitting_subheader_odd', 'fitting_columns_even', 'fitting_columns_odd',
            'fitting_descriptors_even', 'fitting_descriptors_odd'
        The values for each key must be a dictionary, with keys in this internal
        dictionary representing keyword arguments for openpyxl's NamedStyle or
        openpyxl.styles.NamedStyle objects. See
        mcetl.excel_writer.ExcelWriterHandler.styles as an example for the dictionary.

    Returns
    -------
    fit_results : list(lmfit.models.ModelResult)
        A list of lmfit.ModelResult objects, which give information
        for each of the fits done on the dataframes.
    gui_values : dict, optional
        A dictionary containing the default gui values to pass to fit_dataframe.
    proceed : bool
        True if the fitting gui was not exited from prematurely, otherwise,
        the value is False. Useful when calling this function from an
        outside function that needs to know whether to continue doing
        peak fitting.

    """

    rc_params = mpl_changes.copy() if mpl_changes is not None else {}
    # Correctly scales the dpi to match the desired dpi.
    dpi = float(rc_params.get('figure.dpi', plt.rcParams['figure.dpi']))
    rc_params.update({
        'interactive': False,
        'figure.constrained_layout.use': False,
        'figure.autolayout': True,
        'figure.dpi': dpi * plot_utils.get_dpi_correction(dpi)
    })

    if dataframe is not None:
        if isinstance(dataframe, pd.DataFrame):
            fit_dataframes = [dataframe]
        else:
            fit_dataframes = dataframe
    else:
        fit_dataframes = utils.open_multiple_files()

    if save_excel and fit_dataframes:
        if excel_writer is not None:
            writer_handler = ExcelWriterHandler(writer=excel_writer, styles=excel_formats)
        else:
            layout = [
                [sg.Text('File name for peak fitting results')],
                [sg.Input('', key='file', size=(20, 1), disabled=True),
                sg.FileSaveAs(file_types=(("Excel Workbook (xlsx)", "*.xlsx"),),
                            key='browse', target='file')],
                [sg.Text('')],
                [sg.Button('Skip Saving'),
                sg.Button('Submit', bind_return_key=True,
                        button_color=utils.PROCEED_COLOR),
                sg.Check('New File', key='new_file')]
            ]

            window = sg.Window('File Selection', layout, icon=utils._LOGO)
            while True:
                event, values = window.read()
                if event == sg.WIN_CLOSED:
                    utils.safely_close_window(window)
                elif event == 'Skip Saving':
                    save_excel = False
                    break
                elif (event == 'Submit'
                        and utils.validate_inputs(values, strings=[['file', 'File name']])):
                    break

            window.close()
            del window

            if save_excel:
                file_path = Path(values['file'])
                if not file_path.suffix or file_path.suffix.lower() != '.xlsx':
                    values['file'] = str(Path(file_path.parent, file_path.stem + '.xlsx'))

                writer_handler = ExcelWriterHandler(values['file'], values['new_file'], excel_formats)

    fit_results = []
    proceed = True
    for dataframe in fit_dataframes:
        try:
            with plt.rc_context(rc_params):
                fit_output = fit_dataframe(dataframe, gui_values)

        except (utils.WindowCloseError, KeyboardInterrupt):
            proceed = False
            break

        if fit_output[4] is None: # Fitting was skipped for the data entry
            fit_results.append(None)
        else:
            fit_results.append(fit_output[0])
            peak_df = fit_output[1]
            params_df = fit_output[2]
            descriptors_df = fit_output[3]
            gui_values = fit_output[4]

            if save_excel:
                fit_to_excel(peak_df, params_df, descriptors_df,
                             writer_handler, gui_values['sample_name'], plot_excel)

    if save_excel and save_when_done and not all(entry is None for entry in fit_results):
        writer_handler.save_excel_file()

    return fit_results, gui_values, proceed
