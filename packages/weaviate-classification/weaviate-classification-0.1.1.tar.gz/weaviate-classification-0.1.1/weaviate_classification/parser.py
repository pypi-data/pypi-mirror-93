""" This module parses datapoints from an excel file """

import fnmatch
import openpyxl
from .utilities import get_reference_class_name_from_field


def _initialize_result(dataconfig):

    # Initialize the return value
    result = {}
    result['datapoints'] = []
    result['count'] = 0
    result['entities'] = {}

    # Collect all entities from the data configuration
    for dataclass in dataconfig['classes']:
        for col in dataclass['columns']:
            if col['entity']:
                classname = get_reference_class_name_from_field(col['name'])
                result['entities'][classname] = []

    return result


def _parse_data_line(dataconfig, data, sheet, row, dataclass):

    # initialize the return value
    entry = {}

    # store the class name of this entry
    entry['classname'] = dataclass['classname']

    # store the row number of this data point
    entry['row'] = row

    # indicate whether this is a validated data point and (thus) can be used for training. Default is False
    entry['validated'] = False

    # set the batch number for this item will be set at import if batching is used
    entry['batchNumber'] = 1

    # for each column in the data, read the appropriate value from the excel sheet
    for col in dataclass['columns']:

        # cast the value read from the excel sheet to the appropriate type
        if col['type'] == 'int':
            if sheet.cell(row=row, column=col['number']).value is not None:
                value = int(sheet.cell(row=row, column=col['number']).value)
            else:
                value = 0

        elif col['type'] == 'string':
            if sheet.cell(row=row, column=col['number']).value is not None:
                value = str(sheet.cell(row=row, column=col['number']).value)
            else:
                value = ""

        elif col['type'] == 'text':
            if sheet.cell(row=row, column=col['number']).value is not None:
                value = str(sheet.cell(row=row, column=col['number']).value)
            else:
                value = ""

        elif col['type'] == 'date':
            if sheet.cell(row=row, column=col['number']).value is not None:
                value = str(sheet.cell(row=row, column=col['number']).value)
            else:
                value = ""

        elif col['type'] == 'geoCoordinates':
            if sheet.cell(row=row, column=col['number']).value is not None:
                value = str(sheet.cell(row=row, column=col['number']).value)
            else:
                value = ""

        elif col['type'] == 'phoneNumber':
            if sheet.cell(row=row, column=col['number']).value is not None:
                value = str(sheet.cell(row=row, column=col['number']).value)
            else:
                value = ""

        elif col['type'] == 'number':
            if sheet.cell(row=row, column=col['number']).value is not None:
                value = float(sheet.cell(row=row, column=col['number']).value)
            else:
                value = ""

        elif col['type'] == 'boolean':
            if sheet.cell(row=row, column=col['number']).value is not None:
                value = bool(sheet.cell(row=row, column=col['number']).value)
            else:
                value = False

        # Store the value of the property with the data entry point
        entry[col['name']] = value

        # for each entity check if the value is already in the list of values for this entity
        if col['entity']:
            classname = get_reference_class_name_from_field(col['name'])
            if value not in data['entities'][classname]:
                data['entities'][classname].append(value)

    return entry


def _parse_data_sheet(dataconfig, sheet):

    # Initialize the return value
    data = _initialize_result(dataconfig)

    # process the data line by line
    end = False
    row = 2
    while not end:
        source = sheet.cell(row=row, column=1).value
        if source is None:
            end = True
        else:
            for dataclass in dataconfig['classes']:
                point = _parse_data_line(dataconfig, data, sheet, row, dataclass)
                data['datapoints'].append(point)
            data['count'] += 1
            row += 1

    return data


def parse_data_file(dataconfig, filename):
    """ parse the data file """

    # Initialize the return value
    data = None

    # Read the path to the data file from the config file
    if filename is not None:

        # Open the file - assume there is only one sheet with the data in the file
        if fnmatch.fnmatch(filename, "*.xls") or fnmatch.fnmatch(filename, "*.xlsx"):
            workbook = openpyxl.load_workbook(filename, data_only=True)
            if workbook is not None:
                sheet = workbook.active
                if sheet is not None:
                    data = _parse_data_sheet(dataconfig, sheet)

    return data
