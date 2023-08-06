from __future__ import absolute_import
from __future__ import unicode_literals

import datetime as dt
from decimal import Decimal
from io import BytesIO
from random import randint
import os.path as osp

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import xlwt
except ImportError:
    xlwt = None

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None


from .decorators import deprecate


def _xlrd_required():
    if xlrd is None:
        raise ImportError('xlrd library is required to use this function or class')


def _xlwt_required():
    if xlwt is None:
        raise ImportError('xlwt library is required to use this function or class')


def http_headers(filename, randomize=True):
    basename, ext = osp.splitext(filename)
    if randomize:
        rand_filename = '{}-{}{}'.format(basename, randint(1000000, 9999999), ext)
        headers = {'Content-Disposition': 'attachment; filename={}'.format(rand_filename)}
    else:
        headers = {'Content-Disposition': 'attachment; filename={}'.format(filename)}

    if ext == '.xlsx':
        headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument' \
                                  '.spreadsheetml.sheet'
    elif ext == '.xls':
        headers['Content-Type'] = 'application/vnd.ms-excel'
    else:
        raise ValueError('filename "{}" does not end with .xls or .xlsx'.format(filename))

    return headers


@deprecate('xlrd is no longer maintained, recommend switching to openpyxl')
def workbook_to_reader(xlwt_wb):
    """
        convert xlwt Workbook instance to an xlrd instance for reading
    """
    _xlrd_required()
    fh = BytesIO()
    xlwt_wb.save(fh)
    # prep for reading
    fh.seek(0)
    return xlrd.open_workbook(file_contents=fh.read())


def xlsx_to_strio(xlsx_wb):
    """
        convert xlwt Workbook instance to a BytesIO instance
    """
    fh = BytesIO()
    xlsx_wb.filename = fh
    xlsx_wb.close()
    # prep for reading
    fh.seek(0)
    return fh


@deprecate('xlrd is no longer maintained, recommend switching to openpyxl')
def xlsx_to_reader(xlsx_wb):
    """
        convert xlsxwriter Workbook instance to an xlrd instance for reading
    """
    _xlrd_required()
    fh = xlsx_to_strio(xlsx_wb)
    return xlrd.open_workbook(file_contents=fh.read())


class Writer(object):
    STYLE_FACTORY = {}
    FONT_FACTORY = {}

    def __init__(self, ws=None):
        _xlwt_required()
        self.ws = ws
        self.rownum = 0
        self.colnum = 0

    def set_sheet(self, ws):
        self.ws = ws
        self.rownum = 0
        self.colnum = 0

    def write(self, row, col, data, style=None):
        """
        Write data to row, col of worksheet (ws) using the style
        information.

        Again, I'm wrapping this because you'll have to do it if you
        create large amounts of formatted entries in your spreadsheet
        (else Excel, but probably not OOo will crash).
        """
        ws = self.ws
        if not ws:
            raise Exception('you must use set_sheet() before write()')

        if style:
            if isinstance(style, xlwt.Style.XFStyle):
                s = style
            else:
                s = self.get_style(style)
            ws.write(row, col, data, s)
        else:
            ws.write(row, col, data)

    def write_merge(self, r1, r2, c1, c2, data, style=None):
        """
        Write data to row, col of worksheet (ws) using the style
        information.

        Again, I'm wrapping this because you'll have to do it if you
        create large amounts of formatted entries in your spreadsheet
        (else Excel, but probably not OOo will crash).
        """
        ws = self.ws
        if not ws:
            raise Exception('you must use set_sheet() before write()')

        if style:
            if isinstance(style, xlwt.Style.XFStyle):
                s = style
            else:
                s = self.get_style(style)
            ws.write_merge(r1, r2, c1, c2, data, s)
        else:
            ws.write_merge(r1, r2, c1, c2, data)

    def mwrite(self, col_vals, style=None, nextrow=False):
        for val in col_vals:
            self.awrite(val, style)
        if nextrow:
            self.newrow()

    def awrite(self, data=None, style=None, nextrow=False):
        """
            Auto Write: Similar to write, except that the row and column
            numbers are handled automatically and based on the extra
            parameters to this method.
        """
        self.write(self.rownum, self.colnum, data, style)
        self.colnum += 1
        if nextrow:
            self.newrow()

    def newrow(self):
        self.rownum += 1
        self.colnum = 0

    def get_style(self, style):
        """
        Style is a dict maping key to values.
        Valid keys are: background, format, alignment, border

        The values for keys are lists of tuples containing (attribute,
        value) pairs to set on model instances...
        """
        style_key = tuple(style.items())
        s = self.STYLE_FACTORY.get(style_key, None)
        if s is None:
            s = xlwt.XFStyle()
            for key, values in style.items():
                if key == "background":
                    p = xlwt.Pattern()
                    for attr, value in values:
                        p.__setattr__(attr, value)
                    s.pattern = p
                elif key == "format":
                    s.num_format_str = values
                elif key == "alignment":
                    a = xlwt.Alignment()
                    for attr, value in values:
                        a.__setattr__(attr, value)
                    s.alignment = a
                elif key == "border":
                    b = xlwt.Formatting.Borders()
                    for attr, value in values:
                        b.__setattr__(attr, value)
                    s.borders = b
                elif key == "font":
                    f = self.get_font(values)
                    s.font = f
            self.STYLE_FACTORY[style_key] = s
        return s

    def get_font(self, values):
        """
        'height' 10pt = 200, 8pt = 160
        """
        font_key = values
        f = self.FONT_FACTORY.get(font_key, None)
        if f is None:
            f = xlwt.Font()
            for attr, value in values:
                f.__setattr__(attr, value)
            self.FONT_FACTORY[font_key] = f
        return f


class _OpenpyxlWriter:
    def __init__(self, ws):
        self.set_sheet(ws)

    def set_sheet(self, ws):
        self.ws = ws
        # Openpyxl uses 1-based indexing for cells
        self.rownum = 1
        self.colnum = 1

    def awrite(self, data, style=None, nextrow=False):
        cell = self.ws.cell(column=self.colnum, row=self.rownum, value=data)
        if style:
            cell.style = style
        self.colnum += 1
        if nextrow:
            self.nextrow()

    def mwrite(self, *colvals, style=None, nextrow=False):
        for val in colvals:
            self.awrite(val, style)
        if nextrow:
            self.nextrow()

    def nextrow(self):
        self.rownum += 1
        self.colnum = 1


class _XlsxwriterWriter:

    def __init__(self, ws):
        self.set_sheet(ws)

    def set_sheet(self, ws):
        self.ws = ws
        self.rownum = 0
        self.colnum = 0

    def awrite(self, data, style=None, nextrow=False):
        self.ws.write(self.rownum, self.colnum, data, style)
        self.colnum += 1
        if nextrow:
            self.nextrow()

    def mwrite(self, *colvals, style=None, nextrow=False):
        for val in colvals:
            self.awrite(val, style)
        if nextrow:
            self.nextrow()

    def nextrow(self):
        self.rownum += 1
        self.colnum = 0


class WriterX:
    @property
    def ws(self):
        return self._writer.ws

    @property
    def rownum(self):
        return self._writer.rownum

    @rownum.setter
    def rownum(self, value):
        self._writer.rownum = value

    @property
    def colnum(self):
        return self._writer.colnum

    @colnum.setter
    def colnum(self, value):
        self._writer.colnum = value

    def __init__(self, ws):
        self._writer = None
        if openpyxl:
            from openpyxl.worksheet.worksheet import Worksheet
            if isinstance(ws, Worksheet):
                self._writer = _OpenpyxlWriter(ws)
        if xlsxwriter:
            from xlsxwriter.worksheet import Worksheet
            if isinstance(ws, Worksheet):
                self._writer = _XlsxwriterWriter(ws)

        if not self._writer:
            raise TypeError('worksheet type not supported')

    def set_sheet(self, ws):
        self._writer.set_sheet(ws)

    def awrite(self, data, style=None, nextrow=False):
        self._writer.awrite(data, style=style, nextrow=nextrow)

    def mwrite(self, *colvals, style=None, nextrow=False):
        self._writer.mwrite(*colvals, style=style, nextrow=nextrow)

    def nextrow(self):
        self._writer.nextrow()


class XlwtHelper(Writer):

    @deprecate('XlwtHelper has been renamed to Writer')
    def __init__(self, ws=None):
        Writer.__init__(self, ws)


class _XlrdReader:
    @deprecate('xlrd is no longer maintained, recommend switching to openpyxl')
    def __init__(self, xlrd_wb, sheetnum=0):
        self.book = xlrd_wb
        self.sheetnum = sheetnum
        self.rownum = 0
        self.colnum = 0
        self.sheet = self.book.sheet_by_index(self.sheetnum)

    def cell_value(self, is_date=False):
        try:
            return self.sheet.cell_value(self.rownum, self.colnum)
        finally:
            self.colnum += 1

    def cell_date(self):
        value = self.cell_value()
        date_tuple = xlrd.xldate_as_tuple(value, self.book.datemode)
        return dt.datetime(*date_tuple).date()

    def cell_datetime(self):
        value = self.cell_value()
        date_tuple = xlrd.xldate_as_tuple(value, self.book.datemode)
        return dt.datetime(*date_tuple)

    def cell_decimal(self):
        value = self.cell_value()
        return Decimal(value)

    def next_row(self):
        self.rownum += 1
        self.colnum = 0


class _OpenpyxlReader:
    def __init__(self, openpyxl_wb, sheetnum=0):
        self.book = openpyxl_wb
        self.sheetnum = sheetnum
        # Openpyxl uses 1-based indexing for cells
        self.rownum = 1
        self.colnum = 1
        self.sheet = self.book[self.book.sheetnames[sheetnum]]

    def cell_value(self, is_date=False):
        try:
            return self.sheet.cell(self.rownum, self.colnum).value
        finally:
            self.colnum += 1

    def cell_date(self):
        # Openpyxl automatically converts to date or datetime
        value = self.cell_value()
        if isinstance(value, dt.datetime):
            return value.date()
        elif isinstance(value, dt.date):
            return value
        else:
            raise TypeError(f'{value}: not a date')

    def cell_datetime(self):
        # Openpyxl automatically converts to date or datetime
        value = self.cell_value()
        if isinstance(value, dt.datetime):
            return value
        elif isinstance(value, dt.date):
            return dt.datetime(value.year, value.month, value.day)
        else:
            raise ValueError(f'{value}: not a datetime')

    def cell_decimal(self):
        value = self.cell_value()
        return Decimal(value)

    def next_row(self):
        self.rownum += 1
        self.colnum = 1


class Reader(object):
    def __init__(self, workbook, sheetnum=0):
        self._reader = None
        if openpyxl:
            from openpyxl.workbook import Workbook
            if isinstance(workbook, Workbook):
                self._reader = _OpenpyxlReader(workbook, sheetnum=sheetnum)
        if xlrd:
            if xlsxwriter:
                from xlsxwriter import Workbook
                if isinstance(workbook, Workbook):
                    self._reader = _XlrdReader(workbook, sheetnum=sheetnum)
            if xlwt:
                from xlwt.Workbook import Workbook
                if isinstance(workbook, Workbook):
                    self._reader = _XlrdReader(workbook, sheetnum=sheetnum)

        if not self._reader:
            raise TypeError('workbook type not supported')

    @classmethod
    def from_xlwt(cls, xlwt_wb):
        wb = workbook_to_reader(xlwt_wb)
        return _XlrdReader(wb)

    @classmethod
    def from_xlsx(cls, xlsx_wb):
        wb = xlsx_to_reader(xlsx_wb)
        return _XlrdReader(wb)

    def cell_value(self, is_date=False):
        return self._reader.cell_value(is_date=is_date)

    def cell_date(self):
        return self._reader.cell_date()

    def cell_datetime(self):
        return self._reader.cell_datetime()

    def cell_decimal(self):
        return self._reader.cell_decimal()

    def next_row(self):
        self._reader.next_row()
