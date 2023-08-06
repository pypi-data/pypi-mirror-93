import xlrd
import openpyxl


class Read(object):
    def __init__(self, mark_excel):
        self.mark_excel = mark_excel

    def read(self, local_path):
        """
        读取EXCEl文件
        :param local_path: 文件路径
        :return: 返回读取数据列表
        """
        if not local_path.endswith((".xls", ".xlsx")):
            raise Exception("只支持.xls、.xlsx后缀的Excel文件")
        if local_path.endswith(".xls"):
            return self._read_xls(local_path)
        return self._read_xlsx(local_path)

    def _read_xls(self, local_path):
        # 读取excel
        # 打开Excel文件读取数据
        data = xlrd.open_workbook(local_path)
        # 获取第一个工作表
        table = data.sheet_by_index(0)
        # 获取行数
        row_num = table.nrows
        # 获取列数
        col_num = table.ncols
        # 定义excel_list
        excel_list = []
        for row in range(0, row_num):
            item = []
            for col in range(col_num):
                # 获取单元格数据
                cell_value = table.cell(row, col).value
                # 把数据追加到excel_list中
                item.append(cell_value)
            excel_list.append(item)
        return excel_list

    def _read_xlsx(self, local_path):
        # 读取excel
        # 打开Excel文件读取数据
        data = openpyxl.load_workbook(local_path)
        # 获取第一个工作表
        sheet = data[data.sheetnames[0]]
        # 定义excel_list
        excel_list = []
        for row in sheet.rows:
            item = []
            if row[0].value is None:
                continue
            for cell in row:
                # 把数据追加到excel_list中
                item.append(cell.value)
            excel_list.append(item)
        return excel_list
